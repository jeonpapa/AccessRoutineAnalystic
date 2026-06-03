"""
대쉬보드 검색 API 서버 (Flask)
- 대쉬보드에서 약제명 검색 시 해외 약가 실시간 조회
- 국내 약가 DB 검색도 제공
- 국내 약가 변동 이력 및 변동 사유 제공
- 로컬 전용 (127.0.0.1)
"""

import asyncio
import logging
import re
import sqlite3
import sys
from pathlib import Path

from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS

# 프로젝트 루트를 sys.path에 추가
BASE_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(BASE_DIR))

from agents.db import DrugPriceDB
from agents.db.users import UsersDB
from agents.foreign_price_agent import ForeignPriceAgent, AVAILABLE_COUNTRIES
from agents.market_intelligence import MarketIntelligenceAgent, MI_RULES_TEXT
from agents.review_agent import ReviewAgent
from agents.drug_enrichment_agent import DrugEnrichmentAgent
from api.auth import build_auth_blueprint, require_auth
from agents.ingest.market_share import ingest as ingest_market_share
from agents import media_intelligence as _media_intel

app = Flask(__name__)
CORS(
    app,
    resources={r"/api/*": {"origins": ["http://localhost:3000", "http://127.0.0.1:3000"]}},
    supports_credentials=False,
)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

db = DrugPriceDB(BASE_DIR / "data" / "db" / "drug_prices.db")
foreign_agent = ForeignPriceAgent(BASE_DIR)
users_db = UsersDB(BASE_DIR / "data" / "db" / "users.db")
app.register_blueprint(build_auth_blueprint(users_db))


def _check_calibration_age() -> None:
    """서버 시작 시 마지막 캘리브레이션 경과일 확인. 90일 초과면 경고."""
    try:
        from datetime import datetime
        from agents.media_calibrator import load_latest_calibration
        cal = load_latest_calibration()
        if cal is None:
            logger.warning(
                "[MediaCalibrator] 초기 보정 미실행. "
                "POST /api/admin/calibrate-media 또는 "
                "bash scripts/run_calibration.sh 를 실행하세요."
            )
            return
        cal_dt   = datetime.fromisoformat(cal["calibrated_at"])
        days_ago = (datetime.now() - cal_dt).days
        if days_ago >= 90:
            logger.warning(
                "[MediaCalibrator] 마지막 보정 %d일 전 (%s) — 분기 재보정 권장",
                days_ago, cal["calibrated_at"][:10]
            )
        else:
            logger.info("[MediaCalibrator] 최근 보정 %s (%d일 전)", cal["calibrated_at"][:10], days_ago)
    except Exception as e:
        logger.debug("[MediaCalibrator] 상태 확인 실패: %s", e)


_check_calibration_age()


# ──────────────────────────────────────────────────────────────────────────────
# 국내 약가 검색
# ──────────────────────────────────────────────────────────────────────────────

@app.get("/api/domestic/search")
def domestic_search():
    """
    국내 약가 검색
    GET /api/domestic/search?q=키트루다&limit=20
    """
    query = request.args.get("q", "").strip()
    limit = min(int(request.args.get("limit", 50)), 200)
    if not query:
        return jsonify({"error": "검색어(q)를 입력하세요."}), 400

    results = db.search_drug(query, limit=limit)
    return jsonify({"query": query, "count": len(results), "results": results})


@app.get("/api/domestic/history/<insurance_code>")
def domestic_history(insurance_code: str):
    """
    보험코드별 국내 약가 이력
    GET /api/domestic/history/652902770
    """
    results = db.get_price_history(insurance_code)
    return jsonify({"insurance_code": insurance_code, "count": len(results), "results": results})


@app.get("/api/domestic/stats")
def domestic_stats():
    return jsonify(db.get_stats())


# ──────────────────────────────────────────────────────────────────────────────
# 국내 약가 변동 이력
# ──────────────────────────────────────────────────────────────────────────────

_NAME_RE = re.compile(r"^([^(（]+)")          # 브랜드명
_ING_RE  = re.compile(r"\(([^)]+)\)")         # 첫 번째 괄호 = 주성분
_DOSE_RE = re.compile(r"_\(([^)]+)\)\s*$")   # 말미 _(...) = 제형


def _normalize_brand(name: str) -> str:
    """
    제품명을 정규화하여 병합 키로 사용.
    예:
      '자누비아정100mg'                                              → '자누비아정100mg'
      '자누비아정100밀리그램(인산시타글립틴일수화물)'                 → '자누비아정100mg'
      '자누비아정100밀리그램(시타글립틴인산염수화물)_(0.1289g/1정)'    → '자누비아정100mg'
      '리피토정10밀리그람(아토르바스타틴칼슘)'                         → '리피토정10mg'

    규칙:
      1) 뒤쪽 _(...) 제형/함량 suffix 제거
      2) 남은 괄호(...) 성분명 변형 제거
      3) 용량 단위 통일: 밀리그램/밀리그람 → mg, 마이크로그램 → mcg,
         밀리리터 → ml, 그램(단독) → g
      4) 공백·후행 underscore 제거
    """
    if not name:
        return ""
    s = re.sub(r"_\([^)]*\)\s*$", "", name)      # 제형 suffix
    s = re.sub(r"\([^)]*\)", "", s)              # 성분 괄호
    s = s.rstrip("_").strip()
    s = s.replace("밀리그램", "mg").replace("밀리그람", "mg")
    s = s.replace("마이크로그램", "mcg").replace("마이크로그람", "mcg")
    s = s.replace("밀리리터", "ml").replace("밀리리타", "ml")
    # '그램'은 '밀리그램/마이크로그램' 치환 후에만 단독 g 처리
    s = re.sub(r"(\d+)\s*그램", r"\1g", s)
    # 공백 전체 제거 (HIRA는 공백 변형 많음)
    s = re.sub(r"\s+", "", s)
    return s.strip()


def _parse_product(name: str) -> dict:
    """product_name_kr 에서 브랜드명·주성분·제형을 추출."""
    brand_m = _NAME_RE.search(name)
    brand   = brand_m.group(1).strip() if brand_m else name

    # 괄호 목록 전부 추출 → 첫 번째가 주성분 (단, 숫자만이면 농도표기이므로 건너뜀)
    all_parens = re.findall(r"\(([^)]+)\)", name)
    ingredient = ""
    for p in all_parens:
        if not re.match(r"^[\d%\.]+$", p):   # 순수 숫자/% 아닌 것
            ingredient = p
            break

    dose_m  = _DOSE_RE.search(name)
    dosage_form = dose_m.group(1) if dose_m else ""

    return {"brand": brand, "ingredient": ingredient, "dosage_form": dosage_form}


def _extract_dose_unit(product_name: str) -> str:
    """product_name_kr 에서 제형 단위를 추출.
    예) '키트루다주' → '주', '글리벡정' → '정', '자누비아캡슐' → '캡슐'.
    실패 시 빈 문자열 반환.
    """
    if not product_name:
        return ""
    m = re.search(r"(정|주|캡슐|시럽|액|연고|겔|패취|산|과립|분무제|분말|필름|좌제)", product_name)
    return m.group(1) if m else ""


def _build_price_changes(rows: list) -> list:
    """
    DB rows(apply_date 순 정렬) → 가격이 바뀐 시점만 추출.
    반환: [{"date", "price", "price_change", "delta_pct",
           "base_price_change_rate", "change_type", "is_first"}, ...]
      - price_change            : 직전 대비 절대 변동액 (원). 최초는 0.
      - delta_pct               : 직전 대비 변동률 (%). 최초는 None.
      - base_price_change_rate  : 최초 등재가 대비 누적 변동률 (%). 최초는 0.
      - change_type             : '최초' / '인상' / '인하'
    """
    changes = []
    prev_price = None
    base_price = None
    for row in rows:
        price = row["max_price"]
        if price is None:
            continue
        if prev_price is None:
            base_price = price
            changes.append({
                "date": row["apply_date"],
                "price": price,
                "price_change": 0,
                "delta_pct": None,
                "base_price_change_rate": 0.0,
                "change_type": "최초",
                "is_first": True,
            })
        elif price != prev_price:
            abs_delta = price - prev_price
            delta_pct = round(abs_delta / prev_price * 100, 2)
            base_rate = round((price - base_price) / base_price * 100, 2) if base_price else 0.0
            changes.append({
                "date": row["apply_date"],
                "price": price,
                "price_change": abs_delta,
                "delta_pct": delta_pct,
                "base_price_change_rate": base_rate,
                "change_type": "인상" if abs_delta > 0 else "인하",
                "is_first": False,
            })
        prev_price = price
    return changes


@app.get("/api/domestic/price-changes")
def price_changes():
    """
    약제명·주성분명으로 검색 → 보험코드별 가격 변동 이력 반환.
    GET /api/domestic/price-changes?q=키트루다

    반환 구조:
    {
      "query": "...",
      "products": [
        {
          "insurance_code": "...",
          "product_name": "...",
          "brand_name": "...",
          "ingredient": "...",
          "dosage_form": "...",
          "company": "...",
          "first_date": "...",
          "current_price": 0,
          "price_history": [{"date","price","delta_pct","is_first"}, ...]
        }
      ],
      "dosage_forms": ["0.1g/4mL", ...]   // 복수일 때만 필터 활성화
    }
    """
    query = request.args.get("q", "").strip()
    if not query:
        return jsonify({"error": "검색어(q)를 입력하세요."}), 400

    # 1) 키워드로 최신 레코드 검색 → 보험코드 목록 확보
    matches = db.search_drug(query, limit=200)
    if not matches:
        return jsonify({"query": query, "products": [], "dosage_forms": []})

    # 보험코드 중복 제거 (최신 레코드 기준)
    code_map: dict = {}
    for m in matches:
        code = m["insurance_code"]
        if code not in code_map or m["apply_date"] > code_map[code]["apply_date"]:
            code_map[code] = m

    # 2) 보험코드별 상품 조립
    raw_products = []
    for code, latest in code_map.items():
        history_rows = db.get_price_history(code)
        if not history_rows:
            continue

        parsed = _parse_product(latest["product_name_kr"])
        changes = _build_price_changes(history_rows)
        if not changes:
            continue

        raw_products.append({
            "insurance_code": code,
            "product_name": latest["product_name_kr"],
            "brand_name": parsed["brand"],
            "ingredient": parsed["ingredient"],
            # HIRA 영문 `주성분명` 원문 — generic-key 매칭용 (enrichment 상속 경로)
            "hira_ingredient": latest.get("ingredient") or "",
            "dosage_strength": latest.get("dosage_strength") or "",
            "dosage_form": parsed["dosage_form"],
            "company": latest["company"],
            "first_date": changes[0]["date"],
            "current_price": changes[-1]["price"],
            "price_history": changes,
            "_apply_date": latest.get("apply_date", ""),
        })

    # 3) 동일 제품 병합 — _normalize_brand() 단일 키로 통합
    #    브랜드명+용량만 일치하면 (회사·코드·성분표기·제형 suffix 무시) 같은 제품으로 간주.
    #    예: '자누비아정100mg' / '자누비아정100밀리그램(인산시타글립틴일수화물)' /
    #        '자누비아정100밀리그램(시타글립틴인산염수화물)_(0.1289g/1정)' → 모두 '자누비아정100mg'
    merge_map: dict = {}
    for rp in raw_products:
        norm_key = _normalize_brand(rp["product_name"])
        if not norm_key:  # 정규화 실패 — 원본 코드로 개별 유지
            merge_map[f"__single__::{rp['insurance_code']}"] = rp
            continue
        if norm_key not in merge_map:
            rp["normalized_name"] = norm_key
            merge_map[norm_key] = rp
            continue
        # 병합
        existing = merge_map[norm_key]
        by_date: dict = {h["date"]: h for h in existing["price_history"]}
        for h in rp["price_history"]:
            prev = by_date.get(h["date"])
            if not prev or (h.get("price") or 0) > (prev.get("price") or 0):
                by_date[h["date"]] = h
        merged_hist = sorted(by_date.values(), key=lambda x: x["date"])
        synth_rows = [{"apply_date": h["date"], "max_price": h["price"]} for h in merged_hist]
        existing["price_history"] = _build_price_changes(synth_rows)
        existing["first_date"]    = existing["price_history"][0]["date"]
        existing["current_price"] = existing["price_history"][-1]["price"]
        # 대표 메타 = 최신 apply_date 레코드
        if rp["_apply_date"] > existing.get("_apply_date", ""):
            existing["insurance_code"] = rp["insurance_code"]
            existing["product_name"]   = rp["product_name"]
            existing["company"]        = rp["company"]
            existing["ingredient"]     = rp["ingredient"] or existing["ingredient"]
            existing["dosage_form"]    = rp["dosage_form"] or existing["dosage_form"]
            existing["_apply_date"]    = rp["_apply_date"]
        # 이력 메타
        existing.setdefault("merged_codes", [])
        if rp["insurance_code"] not in existing["merged_codes"]:
            existing["merged_codes"].append(rp["insurance_code"])
        existing.setdefault("merged_companies", [])
        if rp["company"] and rp["company"] not in existing["merged_companies"]:
            existing["merged_companies"].append(rp["company"])

    products = list(merge_map.values())

    # 4) 대표 코드도 merged_codes에 포함되도록 보정 + status 산출
    from datetime import datetime as _dt
    today = _dt.now()
    for p in products:
        p.setdefault("merged_codes", [])
        if p["insurance_code"] not in p["merged_codes"]:
            p["merged_codes"].insert(0, p["insurance_code"])
        p.setdefault("merged_companies", [])
        if p["company"] and p["company"] not in p["merged_companies"]:
            p["merged_companies"].insert(0, p["company"])
        # 최신 가격 상태 판정: None/0 또는 마지막 apply_date > 12개월 경과 시 delisted 의심
        last = p["price_history"][-1] if p["price_history"] else None
        status = "active"
        status_detail = ""
        if last:
            try:
                last_dt = _dt.strptime(last["date"], "%Y.%m.%d")
                gap_days = (today - last_dt).days
            except Exception:
                gap_days = 0
            if last.get("price") in (None, 0):
                status = "delisted_probable"
                status_detail = "최신 레코드에 약가 정보 없음 — 급여기준 삭제 또는 제품명 변경 가능성"
            elif gap_days > 365:
                status = "stale"
                status_detail = f"마지막 약가 갱신 이후 {gap_days//30}개월 경과 — 급여 삭제 가능성 검토 필요"
        p["status"]        = status
        p["status_detail"] = status_detail
        p.pop("_apply_date", None)

    # 5) drug_enrichment LEFT JOIN — 허가일·용법용량·투약비 계산
    _enrich_products(products)

    # 제형 기준 정렬
    products.sort(key=lambda x: x["dosage_form"])

    dosage_forms = sorted({p["dosage_form"] for p in products if p["dosage_form"]})

    return jsonify({
        "query": query,
        "products": products,
        "dosage_forms": dosage_forms,
    })


_GENERIC_FIRST_WORD_RE = re.compile(r"^([a-zA-Z][a-zA-Z\-]+)")
# salt/ester 접미사가 주인공 generic 뒤에 오는 경우가 많아 첫 영단어만으로 매칭 (sitagliptin phosphate hydrate vs sitagliptin hydrochloride hydrate → 모두 'sitagliptin')

_GENERIC_ALL_WORDS_RE = re.compile(r"(?:^|[\s,;])([a-zA-Z][a-zA-Z\-]{3,})")
# 복합제의 모든 영문 generic 후보 추출 (예: 'metformin hydrochloride 0.5g, sitagliptin' → ['metformin', 'hydrochloride', 'sitagliptin'])
# salt/form 단어 (hydrochloride, sulfate, calcium 등) 는 _SALT_WORDS 로 제외
_SALT_WORDS = {
    "hydrochloride", "hydrate", "dihydrate", "trihydrate", "monohydrate", "anhydrous",
    "phosphate", "sulfate", "sulphate", "citrate", "maleate", "fumarate", "tartrate",
    "succinate", "mesylate", "tosylate", "besylate", "acetate", "propanediol",
    "calcium", "sodium", "potassium", "magnesium", "zinc", "iron", "lithium",
    "as", "at", "of", "and", "or", "the", "to", "for", "with", "in", "on",
}
_STRENGTH_MG_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(?:밀리그램|mg|㎎)", re.IGNORECASE)


def _extract_generic_key(ingredient: str | None) -> str:
    """HIRA `주성분명` 첫 영단어 generic 키 — 대표용. _extract_generic_candidates 도 함께 사용."""
    if not ingredient:
        return ""
    m = _GENERIC_FIRST_WORD_RE.match(ingredient.strip())
    if not m:
        return ""
    return m.group(1).strip().lower()


def _extract_generic_candidates(ingredient: str | None) -> list[str]:
    """복합제 대응 — ingredient 에서 가능한 모든 generic 후보를 순서대로 반환.

    예: 'metformin hydrochloride 0.5g, sitagliptin phosphate 50mg' → ['metformin', 'sitagliptin']
        'dapagliflozin propanediol hydrate (as dapagliflozin 10mg), sitagliptin ...' → ['dapagliflozin', 'sitagliptin']
    salt/form/hydrate 단어는 제외. donor 탐색 시 첫 매칭 generic 사용.
    """
    if not ingredient:
        return []
    out: list[str] = []
    seen: set[str] = set()
    for m in _GENERIC_ALL_WORDS_RE.finditer(ingredient):
        w = m.group(1).strip().lower()
        if w in _SALT_WORDS or w in seen:
            continue
        seen.add(w)
        out.append(w)
    return out


def _extract_strength_mg(text: str | None) -> float | None:
    """'자누비아정100mg' / '100밀리그램' → 100.0 (mg 단위)."""
    if not text:
        return None
    m = _STRENGTH_MG_RE.search(text)
    if not m:
        return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


# ────────────────────────────────────────────────────────────────────────────
# 체중·BSA 기반 dosing 표준 환자 기준 (HIRA 약제 평가기준)
#   성인 평균체중 60kg / 체표면적 1.7m² (DuBois 공식 기준 60kg, 165cm)
#   소아 평균체중 40kg
#   DuBois: BSA(cm²) = (W^0.425 × H^0.725) × 0.007184
# ────────────────────────────────────────────────────────────────────────────
_ADULT_WEIGHT_KG = 60.0
_ADULT_BSA_M2 = 1.7
_PEDI_WEIGHT_KG = 40.0

_DOSE_PER_KG_RE = re.compile(
    r"(?:체중\s*1?\s*kg\s*당|/\s*kg|kg\s*당)\s*(\d+(?:\.\d+)?)\s*mg|"
    r"(\d+(?:\.\d+)?)\s*mg\s*/\s*kg",
    re.IGNORECASE,
)
_DOSE_PER_M2_RE = re.compile(
    r"(\d+(?:\.\d+)?)\s*mg\s*/\s*m\s*[²2\^]|"
    r"체표면적\s*m[²2\^]?\s*당\s*(\d+(?:\.\d+)?)\s*mg",
    re.IGNORECASE,
)
_INTERVAL_WEEK_RE = re.compile(
    r"매?\s*(\d+)\s*주\s*(?:마다|간격|에)|q\s*(\d+)\s*w",
    re.IGNORECASE,
)
_INTERVAL_DAY_RE = re.compile(
    r"매?\s*(\d+)\s*일\s*(?:마다|간격|에)",
    re.IGNORECASE,
)


# 용법용량(투여방법) 패턴 — 다음 중 하나 이상 매치되면 진짜 용법 텍스트로 판정.
# 적응증·효능 텍스트 ("절제 불가능 비소세포폐암 환자 치료") 와 구분하기 위함.
_DOSING_PATTERN = re.compile(
    r"(?:1\s*일\s*\d|매\s*\d+\s*주|"          # "1일 N회" / "매 N주마다"
    r"\d+\s*주\s*(?:마다|간격|에)|"           # "N주 간격으로"
    r"\d+\s*일\s*(?:마다|간격|에)|"           # "N일 간격으로"
    r"q\s*\d+\s*[whWH]|"                      # "q3w" / "q12h"
    r"\d+\s*mg\s*/\s*kg|kg\s*당\s*\d+|"       # "mg/kg" / "kg당 mg"
    r"\d+\s*mg\s*/\s*m\s*[²2\^]|"             # "mg/m²"
    r"\d+\s*(?:mg|g)\s*(?:을|를)?\s*"          # "200mg 을" "1200mg 을"
    r"\s*(?:매|q|주|회|일|투여|복용|주사|점적))",
    re.IGNORECASE,
)


def _is_dosing_text(text: str | None) -> bool:
    """텍스트가 '실제 투여방법' 인지 '효능·적응증' 인지 휴리스틱 판정.

    True → 용법용량 패턴 (mg/kg, q3w, 매 N주, 1일 N회 등) 1개 이상 매치
    False → 적응증/효능 서술 (예: '백금 기반 CCRT 후 진행되지 않은 ... 환자 치료')
    """
    if not text or len(text) < 10:
        return False
    return bool(_DOSING_PATTERN.search(text))


def _compute_weight_bsa_daily_cost(
    usage_text: str,
    vial_mg: float,
    current_price: float,
) -> dict | None:
    """체중(mg/kg) · BSA(mg/m²) 기반 dosing → 표준 환자(60kg / 1.7m²) 일일투약비용.

    Returns: {"daily_cost": int, "method": "weight"|"bsa", "rationale": str, ...} or None
    """
    if not usage_text or not vial_mg or vial_mg <= 0 or current_price <= 0:
        return None

    # ── Maintenance dose 우선 — chronic dosing 의 daily_cost 가 의미있음 ──
    # 패턴: "초기 4mg/kg, 이후 3주마다 2mg/kg" / "loading 4, maintenance 2"
    # 모든 mg/kg / mg/m² 매치를 list 로 수집 후 maintenance 우선 선택:
    #   1. "이후" / "유지" / "maintenance" 키워드 뒤의 매치 → 우선
    #   2. 여러 매치 중 가장 작은 값 (loading 은 보통 더 큼)
    #   3. 마지막 매치 (전형적으로 maintenance 가 뒤에 서술)
    all_kg = list(_DOSE_PER_KG_RE.finditer(usage_text))
    all_m2 = list(_DOSE_PER_M2_RE.finditer(usage_text))

    def _pick_maintenance(matches: list[re.Match]) -> tuple[float, str] | None:
        """매치 리스트에서 maintenance dose 추정. Returns (mg, rationale_hint)."""
        if not matches:
            return None
        keys = ("이후", "유지", "maintenance", "지속", "장기")
        for m in matches:
            # "이후" 등 키워드와 같은 문장/근접 위치 확인 (앞 50자 내)
            window_start = max(0, m.start() - 50)
            window = usage_text[window_start:m.start()]
            if any(k in window for k in keys):
                val = next(float(g) for g in m.groups() if g)
                return val, f"maintenance dose ('이후/유지' 키워드 매치)"
        # fallback: 가장 작은 값 (loading > maintenance 일반적)
        candidates = []
        for m in matches:
            v = next(float(g) for g in m.groups() if g)
            candidates.append(v)
        if candidates:
            picked = min(candidates)
            if len(candidates) > 1:
                return picked, f"maintenance 추정 (loading 후보 {sorted(candidates, reverse=True)[0]:.0f} 제외, 최소값 {picked:.0f} 사용)"
            return picked, "단일 매치"
        return None

    weight_pick = _pick_maintenance(all_kg)
    bsa_pick = _pick_maintenance(all_m2)
    if not weight_pick and not bsa_pick:
        return None

    week_m = _INTERVAL_WEEK_RE.search(usage_text)
    day_m = _INTERVAL_DAY_RE.search(usage_text)
    if week_m:
        weeks_grp = next((g for g in week_m.groups() if g), None)
        interval_days = int(weeks_grp) * 7 if weeks_grp else None
    elif day_m:
        interval_days = int(day_m.group(1))
    else:
        return None
    if not interval_days or interval_days <= 0:
        return None

    if weight_pick:
        per_kg_mg, hint = weight_pick
        per_dose_mg = per_kg_mg * _ADULT_WEIGHT_KG
        method = "weight"
        rationale = (
            f"{per_kg_mg}mg/kg × 60kg(성인 평균체중 표준) = {per_dose_mg:.0f}mg/dose, "
            f"q{interval_days // 7}w 간격 → 일평균 {per_dose_mg / interval_days:.2f}mg "
            f"[{hint}]"
        )
    else:
        per_m2_mg, hint = bsa_pick  # type: ignore[misc]
        per_dose_mg = per_m2_mg * _ADULT_BSA_M2
        method = "bsa"
        rationale = (
            f"{per_m2_mg}mg/m² × 1.7m²(DuBois BSA 성인 표준 60kg/165cm) = "
            f"{per_dose_mg:.0f}mg/dose, q{interval_days // 7}w → 일평균 {per_dose_mg / interval_days:.2f}mg "
            f"[{hint}]"
        )

    daily_mg = per_dose_mg / interval_days
    price_per_mg = current_price / vial_mg
    daily_cost = int(round(daily_mg * price_per_mg))
    return {
        "daily_cost": daily_cost,
        "daily_mg": daily_mg,
        "per_dose_mg": per_dose_mg,
        "interval_days": interval_days,
        "vial_mg": vial_mg,
        "method": method,
        "rationale": rationale,
    }


def _enrich_products(products: list[dict]) -> None:
    """각 product 에 drug_enrichment + coverage_start 를 붙인다.

    추가 필드:
      - approval_date      (drug_enrichment.approval_date)
      - usage_text         (drug_enrichment.usage_text)
      - coverage_start     (drug_prices.coverage_start 의 earliest non-null, 없으면 first_date)
      - daily_cost         (int | None, 계산식 아래 참고)
      - monthly_cost       (= daily_cost * 30, 있을 때만)
      - yearly_cost        (= daily_cost * 365, 있을 때만)
      - enrichment_confidence (text)

    daily_cost 계산:
      - schedule='continuous' & daily_dose_units: price * daily_dose_units
      - schedule='cycle' & cycle_days & doses_per_cycle:
          cost_per_cycle = price * doses_per_cycle  → daily = cost_per_cycle / cycle_days
      - 기타: None
    """
    if not products:
        return
    norm_keys = {p.get("normalized_name") for p in products if p.get("normalized_name")}
    if not norm_keys:
        for p in products:
            p.update({"approval_date": None, "usage_text": None, "coverage_start": None,
                      "daily_cost": None, "monthly_cost": None, "yearly_cost": None,
                      "enrichment_confidence": None})
        return

    enrich_map: dict[str, dict] = {}
    try:
        placeholders = ",".join(["?"] * len(norm_keys))
        with db._connect() as conn:
            rows = conn.execute(
                f"SELECT normalized_name, approval_date, usage_text, daily_dose_units, "
                f"       dose_schedule, cycle_days, doses_per_cycle, confidence, "
                f"       is_rsa, rsa_type, rsa_note "
                f"FROM drug_enrichment WHERE normalized_name IN ({placeholders})",
                tuple(norm_keys),
            ).fetchall()
            for r in rows:
                enrich_map[r[0]] = {
                    "approval_date":   r[1],
                    "usage_text":      r[2],
                    "daily_dose_units": r[3],
                    "schedule":        r[4],
                    "cycle_days":      r[5],
                    "doses_per_cycle": r[6],
                    "confidence":      r[7],
                    "is_rsa":          r[8],
                    "rsa_type":        r[9],
                    "rsa_note":        r[10],
                }
            # coverage_start: 보험코드별 earliest non-null
            all_codes = []
            for p in products:
                all_codes.extend(p.get("merged_codes") or [p["insurance_code"]])
            if all_codes:
                cov_placeholders = ",".join(["?"] * len(all_codes))
                cov_rows = conn.execute(
                    f"SELECT insurance_code, MIN(coverage_start) FROM drug_prices "
                    f"WHERE insurance_code IN ({cov_placeholders}) "
                    f"AND coverage_start IS NOT NULL AND coverage_start != '' "
                    f"GROUP BY insurance_code",
                    tuple(all_codes),
                ).fetchall()
                coverage_by_code = {r[0]: r[1] for r in cov_rows}
            else:
                coverage_by_code = {}
    except Exception as e:
        logger.warning("enrichment join 실패: %s", e)
        enrich_map = {}
        coverage_by_code = {}

    # ── 동일 generic 상속 맵 — 복합제 대응을 위해 multi-candidate 등록 ─────────────
    #    enriched 제품의 모든 generic 후보를 donor 맵에 기록 (sitagliptin 외 metformin 등도).
    generic_donor: dict[str, tuple[str, dict]] = {}  # generic → (donor_norm, donor_enrich)
    for p in products:
        norm = p.get("normalized_name")
        if not norm:
            continue
        donor = enrich_map.get(norm)
        if not donor:
            continue
        for gk in _extract_generic_candidates(p.get("hira_ingredient")):
            generic_donor.setdefault(gk, (norm, donor))

    # 전역 donor 보강 — drug_enrichment 에 존재하지만 이번 검색결과 products 에 없는 generic 도 커버.
    # 예: '시타글립틴' 으로 검색 시 자누비아 100mg 는 보통 포함되지만, '메트포르민' 검색 시 combo 에서만
    #     sitagliptin 매칭이 필요할 수 있음. 이 경우를 위해 drug_enrichment 전체를 로드하진 않고
    #     이번 products 범위 내 donor 만 사용 (안전한 기본값).

    for p in products:
        norm = p.get("normalized_name")
        e = enrich_map.get(norm) if norm else None
        donor_norm: str | None = None
        if not e:
            for gk in _extract_generic_candidates(p.get("hira_ingredient")):
                if gk in generic_donor:
                    donor_norm, e = generic_donor[gk]
                    break
        current_price = p.get("current_price") or 0

        # ── 1차 권위 소스: 식약처 의약품 제품 허가정보 API ────────────────────
        # 결과 있으면 approval_date / usage_text 를 MFDS 실측으로 override.
        # drug_enrichment(Perplexity) 는 보조 소스로 dose_schedule/daily_dose_units 등 산출에만 활용.
        mfds_permit = None
        try:
            from agents.scrapers.kr_mfds_permit import lookup_permit
            mfds_query = p.get("product_name") or p.get("brand_name") or ""
            mfds_permit = lookup_permit(
                mfds_query,
                ingredient=p.get("ingredient"),
                edi_code=p.get("insurance_code"),
            )
            if mfds_permit.get("source") == "miss":
                mfds_permit = None
        except Exception as ex:
            logger.debug("MFDS 허가 API 실패 [%s]: %s", p.get("insurance_code"), ex)
            mfds_permit = None

        # 기본은 drug_enrichment 결과 → MFDS permit 가 있으면 우선
        approval_date = e["approval_date"] if e else None
        usage_text = e["usage_text"] if e else None
        if mfds_permit:
            if mfds_permit.get("permit_date"):
                approval_date = mfds_permit["permit_date"]
            if mfds_permit.get("usage_text"):
                usage_text = mfds_permit["usage_text"]

        daily_cost = None
        if e and current_price > 0:
            sched = (e.get("schedule") or "").lower()
            if sched == "continuous" and e.get("daily_dose_units"):
                units = float(e["daily_dose_units"])
                # 동일 generic 상속 시 per-strength 스케일 — 일일 총 mg 보존 (monotherapy 한정)
                # 복합제(제품명에 "/" 포함)는 각 성분 함량이 독립 → 단순 스케일 불가.
                # 복합제 ↔ monotherapy 간 상속은 1 tablet/day 로 가정 (scaled=False 로 표시).
                is_combo = bool(norm and "/" in norm)
                donor_is_combo = bool(donor_norm and "/" in donor_norm)
                scaled = False
                if donor_norm and not is_combo and not donor_is_combo:
                    donor_mg = _extract_strength_mg(donor_norm) or _extract_strength_mg(p.get("dosage_strength"))
                    this_mg = _extract_strength_mg(norm) or _extract_strength_mg(p.get("dosage_strength"))
                    if donor_mg and this_mg and this_mg > 0:
                        units = units * (donor_mg / this_mg)
                        scaled = True
                # 복합제 상속은 unit price 추정 (1정/일). enrichment_source 로 투명화.
                daily_cost = int(round(current_price * units))
                # combo 상속 표시 플래그 (없어도 donor_norm 로 식별 가능하지만 명시적으로)
                if donor_norm and (is_combo or donor_is_combo) and not scaled:
                    p.setdefault("_combo_inherited", True)
            elif sched == "cycle" and e.get("cycle_days") and e.get("doses_per_cycle"):
                # cycle 은 strength 별 바이알 수가 처방 프로토콜에 따라 달라 단순 스케일 부적합
                # → 같은 normalized_name 에 대한 직접 enrichment 있을 때만 산출 (donor 상속 시 skip)
                if not donor_norm:
                    per_cycle = current_price * float(e["doses_per_cycle"])
                    try:
                        daily_cost = int(round(per_cycle / float(e["cycle_days"])))
                    except ZeroDivisionError:
                        daily_cost = None

        # ── usage_text 가 진짜 '투여방법' 인지, 효능·적응증 인지 휴리스틱 판정 ──
        # 임핀지 등 일부 약제는 Perplexity 가 effect_text(효능) 를 usage_text 로 잘못 채움.
        # mg/kg, q3w, 매 N주, 1일 N회 등 dosing 패턴 미포함 시 → 적응증으로 간주.
        usage_is_dosing = _is_dosing_text(usage_text)

        # ── 체중·BSA 기반 dosing 보정 (옵디보·키트루다 등 주사제 mg/kg, mg/m² 패턴) ──
        # 진짜 용법 텍스트일 때만 적용. 적응증 텍스트는 mg/kg 미포함이라 자연 skip 되긴 하지만 명시 분기.
        bsa_calc = None
        if usage_is_dosing and current_price > 0:
            vial_mg = _extract_strength_mg(norm) or _extract_strength_mg(p.get("dosage_strength"))
            if vial_mg:
                bsa_calc = _compute_weight_bsa_daily_cost(usage_text, vial_mg, current_price)
                if bsa_calc:
                    daily_cost = bsa_calc["daily_cost"]
                    p["_bsa_calc"] = bsa_calc  # UI 투명화용

        # 전혀 매칭되지 않은 경우 — heuristic fallback. 단 usage_text 가 적응증 텍스트면
        # daily_cost 산출 자체가 신뢰 불가 → unit price 보존하되 `usage_unverified` 마킹.
        _heuristic_used = False
        if daily_cost is None and current_price > 0:
            daily_cost = current_price
            _heuristic_used = True
        if _heuristic_used and not usage_text:
            usage_text = "용법 미확정 · 1정/일 추정 — 정확한 용법은 조사 필요"

        # 용법 미확정 플래그 — usage_text 가 있는데 dosing 패턴 미포함 (= 적응증 가능성)
        usage_unverified = bool(usage_text) and not usage_is_dosing and not bsa_calc

        monthly_cost = daily_cost * 30 if daily_cost is not None else None
        yearly_cost  = daily_cost * 365 if daily_cost is not None else None

        # coverage_start: merged_codes 중 earliest non-null, 없으면 first_date
        cov = None
        for code in (p.get("merged_codes") or [p["insurance_code"]]):
            if code in coverage_by_code:
                cov = coverage_by_code[code] if cov is None or coverage_by_code[code] < cov else cov
        if not cov:
            cov = p.get("first_date")

        p["approval_date"]          = approval_date
        p["usage_text"]             = usage_text
        p["coverage_start"]         = cov
        p["daily_cost"]             = daily_cost
        p["monthly_cost"]           = monthly_cost
        p["yearly_cost"]            = yearly_cost
        p["enrichment_confidence"]  = e.get("confidence") if e else None
        # RSA(위험분담제) — registry 우선 적용 (skill ground-truth), Perplexity 추정값은 fallback
        from agents.kr_rsa_registry import lookup_rsa
        rsa_reg = lookup_rsa(p.get("brand_name") or p.get("product_name") or "")
        if rsa_reg:
            p["is_rsa"]    = rsa_reg["is_rsa"]
            p["rsa_type"]  = rsa_reg["rsa_type"]
            p["rsa_note"]  = rsa_reg["rsa_note"]
            p["rsa_source"] = rsa_reg["source"]
        else:
            p["is_rsa"]    = e.get("is_rsa") if e else None
            p["rsa_type"]  = e.get("rsa_type") if e else None
            p["rsa_note"]  = (e.get("rsa_note") or "") if e else ""
            p["rsa_source"] = "drug_enrichment (Perplexity)" if e else None
        # 상속 근거 투명화: 공백 값이 아닌 상속 결과임을 UI 가 구분할 수 있도록 표시
        p["enrichment_source"]      = (
            "direct" if (e and not donor_norm) else
            ("inherited_generic:" + donor_norm) if donor_norm else
            "default_heuristic" if _heuristic_used else
            None
        )
        # BSA/체중 기반 dosing 산출 결과 expose (UI 가 표준 환자 기준 명시 가능)
        if bsa_calc:
            p["bsa_calc"] = {
                "daily_cost": bsa_calc["daily_cost"],
                "method":     bsa_calc["method"],
                "rationale":  bsa_calc["rationale"],
                "interval_days": bsa_calc["interval_days"],
                "per_dose_mg":   bsa_calc["per_dose_mg"],
            }
        # 용법 미확정 플래그 — UI 가 daily_cost 신뢰도 낮음을 표시
        p["usage_unverified"] = usage_unverified

        # ── MFDS 허가정보 메타 (UI 노출용) ───────────────────────────────────
        if mfds_permit:
            p["mfds_permit"] = {
                "item_seq":       mfds_permit.get("item_seq"),
                "item_eng_name":  mfds_permit.get("item_eng_name"),
                "permit_holder":  mfds_permit.get("entp_name"),
                "permit_date":    mfds_permit.get("permit_date"),
                "cancel_status":  mfds_permit.get("cancel_status"),
                "etc_otc":        mfds_permit.get("etc_otc"),
                "atc_code":       mfds_permit.get("atc_code"),
                "main_ingr":      mfds_permit.get("main_ingr"),
                "main_ingr_eng":  mfds_permit.get("main_ingr_eng"),
                "material_name":  mfds_permit.get("material_name"),
                "chart":          mfds_permit.get("chart"),
                "pack_unit":      mfds_permit.get("pack_unit"),
                "valid_term":     mfds_permit.get("valid_term"),
                "storage_method": mfds_permit.get("storage_method"),
                "rare_drug_yn":   mfds_permit.get("rare_drug_yn"),
                "newdrug_class":  mfds_permit.get("newdrug_class"),
                "reexam_target":  mfds_permit.get("reexam_target"),
                "reexam_date":    mfds_permit.get("reexam_date"),
                "change_date":    mfds_permit.get("change_date"),
                "permit_kind":    mfds_permit.get("permit_kind"),
                "effect_text":    (mfds_permit.get("effect_text") or "")[:2000],
                "caution_text":   (mfds_permit.get("caution_text") or "")[:2000],
                "ud_doc_url":     mfds_permit.get("ud_doc_url"),
                "ee_doc_url":     mfds_permit.get("ee_doc_url"),
                "nb_doc_url":     mfds_permit.get("nb_doc_url"),
                "source":         mfds_permit.get("source"),
            }
            # enrichment_source 갱신: MFDS 가 우선 소스이면 라벨에 명시
            existing_src = p.get("enrichment_source") or ""
            if existing_src in (None, "", "default_heuristic"):
                p["enrichment_source"] = "mfds_permit"
            elif "mfds" not in existing_src:
                p["enrichment_source"] = f"mfds_permit + {existing_src}"
        else:
            p["mfds_permit"] = None

        # ── 특허 상태 — 1차: MFDS 공공데이터 API (실측), 2차: loe_pattern 가격 history 추정 ──
        # 권위 소스: 식약처 의약품 특허정보 (data.go.kr MdcinPatentInfoService2)
        try:
            from agents.scrapers.kr_mfds_patent import lookup_patent
            # item_name 은 product_name 우선, 부족하면 brand_name 으로 fallback
            mfds_query_name = p.get("product_name") or p.get("brand_name") or ""
            mfds_result = lookup_patent(mfds_query_name, ingredient=p.get("ingredient"))
            mfds_status = mfds_result.get("status")  # '유효' / '만료' / 'unknown'
        except Exception as ex:
            logger.debug("MFDS 특허 API 실패 [%s]: %s", p.get("insurance_code"), ex)
            mfds_result = None
            mfds_status = None

        if mfds_status in ("유효", "만료"):
            p["patent_status"] = mfds_status
            p["patent_expiry_date"] = mfds_result.get("expiry_date")
            p["patent_loe_date_inferred"] = None
            p["patent_source"] = "mfds_api"
            p["patent_source_note"] = (
                f"MFDS 의약품 특허정보 API 실측 — 물질특허 {mfds_result.get('active_substance_count')}건 활성 / "
                f"{mfds_result.get('expired_substance_count')}건 만료 (판정 근거: {mfds_result.get('judgment_basis')})"
            )
            p["patent_substance_patents"] = mfds_result.get("substance_patents", [])
        else:
            # API 데이터 없음 → loe_pattern fallback
            try:
                from agents.market_intelligence.loe_pattern import detect_loe_stage
                history = db.get_price_history(p["insurance_code"])
                history_dates = sorted({h.get("apply_date") for h in history if h.get("apply_date")})
                patent_status = "유효"
                patent_loe_date = None
                for date in history_dates:
                    det = detect_loe_stage(history, date)
                    if det.matched and det.kr_rule == "KR-RULE-009":
                        patent_status = "만료"
                        patent_loe_date = det.anchor_date
                        break
                p["patent_status"] = patent_status
                p["patent_loe_date_inferred"] = patent_loe_date
                p["patent_expiry_date"] = None
                p["patent_source"] = "price_history"
                p["patent_source_note"] = (
                    "가격 history 추정 (KR-RULE-009 stage). MFDS API 미매칭 — "
                    "정확한 만료일은 nedrug.mfds.go.kr/searchPatent 직접 조회"
                )
                p["patent_substance_patents"] = []
            except Exception as ex:
                logger.debug("patent_status fallback 실패 [%s]: %s", p.get("insurance_code"), ex)
                p["patent_status"] = None
                p["patent_expiry_date"] = None
                p["patent_loe_date_inferred"] = None
                p["patent_source"] = None
                p["patent_substance_patents"] = []


# ──────────────────────────────────────────────────────────────────────────────
# 엑셀/CSV 다운로드 (가격 변동 이력 · 약가 정보)
# ──────────────────────────────────────────────────────────────────────────────

def _export_rows(
    rows: list[dict],
    columns: list[tuple[str, str]],
    filename_base: str,
    fmt: str,
):
    """
    rows: list of dicts
    columns: [(key, header_label), ...]  — 출력 순서·헤더명
    fmt: 'csv' | 'xlsx'
    """
    import io
    from flask import Response
    from datetime import datetime
    from urllib.parse import quote

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_base = re.sub(r"[^\w\-]", "_", filename_base)[:60] or "export"

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return jsonify({"error": "openpyxl 미설치 — pip install openpyxl"}), 500
        wb = Workbook()
        ws = wb.active
        ws.title = "export"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1A56DB")
        ws.append([label for _, label in columns])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append([r.get(k, "") for k, _ in columns])
        # 컬럼 폭 자동
        for col_idx, (k, _) in enumerate(columns, start=1):
            max_len = max(
                [len(str(r.get(k, ""))) for r in rows] + [len(columns[col_idx-1][1])]
            ) + 2
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len, 40)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"{safe_base}_{stamp}.xlsx"
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
        )

    # CSV (엑셀 한글 호환 BOM 추가)
    import csv as _csv
    buf = io.StringIO()
    buf.write("\ufeff")
    writer = _csv.writer(buf)
    writer.writerow([label for _, label in columns])
    for r in rows:
        writer.writerow([r.get(k, "") for k, _ in columns])
    fname = f"{safe_base}_{stamp}.csv"
    return Response(
        buf.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


@app.get("/api/domestic/price-changes/export")
def price_changes_export():
    """
    가격 변동 이력을 엑셀/CSV 로 다운로드.
    GET /api/domestic/price-changes/export?q=키트루다&format=xlsx
    format: 'csv' | 'xlsx' (default 'xlsx')
    """
    query = request.args.get("q", "").strip()
    fmt   = (request.args.get("format", "xlsx") or "xlsx").lower()
    if fmt not in ("csv", "xlsx"):
        return jsonify({"error": "format 은 csv 또는 xlsx 여야 합니다."}), 400
    if not query:
        return jsonify({"error": "검색어(q)를 입력하세요."}), 400

    # /api/domestic/price-changes 와 동일 로직 재사용
    with app.test_request_context(f"/api/domestic/price-changes?q={query}"):
        payload = price_changes().get_json()
    products = payload.get("products", [])
    if not products:
        return jsonify({"error": f"'{query}' 검색 결과 없음"}), 404

    rows = []
    for p in products:
        base = {
            "product_name":   p.get("product_name", ""),
            "brand_name":     p.get("brand_name", ""),
            "ingredient":     p.get("ingredient", ""),
            "dosage_form":    p.get("dosage_form", ""),
            "company":        p.get("company", ""),
            "insurance_code": p.get("insurance_code", ""),
            "status":         p.get("status", ""),
        }
        for h in p.get("price_history", []):
            rows.append({
                **base,
                "date":      h.get("date", ""),
                "price":     h.get("price", ""),
                "delta_pct": ("" if h.get("is_first") else h.get("delta_pct", "")),
                "is_first":  "등재" if h.get("is_first") else "변동",
            })

    columns = [
        ("date",           "변동일"),
        ("is_first",       "구분"),
        ("price",          "상한금액(원)"),
        ("delta_pct",      "변동률(%)"),
        ("product_name",   "제품명"),
        ("brand_name",     "브랜드"),
        ("ingredient",     "주성분"),
        ("dosage_form",    "규격"),
        ("company",        "업체명"),
        ("insurance_code", "보험코드"),
        ("status",         "상태"),
    ]
    return _export_rows(rows, columns, f"price_history_{query}", fmt)


@app.get("/api/domestic/search/export")
def domestic_search_export():
    """
    약가 정보(최신 레코드)를 엑셀/CSV 로 다운로드.
    GET /api/domestic/search/export?q=키트루다&format=xlsx
    """
    query = request.args.get("q", "").strip()
    fmt   = (request.args.get("format", "xlsx") or "xlsx").lower()
    if fmt not in ("csv", "xlsx"):
        return jsonify({"error": "format 은 csv 또는 xlsx 여야 합니다."}), 400
    if not query:
        return jsonify({"error": "검색어(q)를 입력하세요."}), 400

    results = db.search_drug(query, limit=500)
    if not results:
        return jsonify({"error": f"'{query}' 검색 결과 없음"}), 404

    # DB 레코드 키 그대로 사용 (존재하는 키만 노출)
    sample = results[0]
    preferred = [
        ("apply_date",      "적용일"),
        ("product_name_kr", "제품명(한글)"),
        ("ingredient_kr",   "주성분"),
        ("company",         "업체명"),
        ("insurance_code",  "보험코드"),
        ("max_price",       "상한금액(원)"),
        ("dosage_form",     "제형/규격"),
        ("atc_code",        "ATC코드"),
        ("remark",          "비고"),
    ]
    columns = [(k, label) for k, label in preferred if k in sample]
    # 추가 키가 있으면 뒤에 붙임
    for k in sample.keys():
        if k not in {c[0] for c in columns}:
            columns.append((k, k))
    return _export_rows(results, columns, f"drug_info_{query}", fmt)


# ──────────────────────────────────────────────────────────────────────────────
# 가격 변동 사유 조회 (MarketIntelligenceAgent — 의학전문지 + 기전 분석)
# ──────────────────────────────────────────────────────────────────────────────

_mi_agent = MarketIntelligenceAgent(
    cache_dir=BASE_DIR / "data" / "dashboard" / "reason_cache"
)
_review_agent = ReviewAgent()
_enrichment_agent = DrugEnrichmentAgent(db)


# ──────────────────────────────────────────────────────────────────────────────
# 약제 부가정보 (RSA · 허가일 · 용법용량 · 치료비)
# ──────────────────────────────────────────────────────────────────────────────

@app.get("/api/domestic/enrichment")
def domestic_enrichment():
    """
    GET /api/domestic/enrichment
        ?normalized_name=자누비아정100mg
        &code=498900030
        &product_name=자누비아정100밀리그램(인산시타글립틴일수화물)
        &ingredient=시타글립틴인산염수화물
        &current_price=866
        &codes=498900030,498900031      (병합 코드 목록)
        &refresh=0
    """
    normalized_name = request.args.get("normalized_name", "").strip()
    code            = request.args.get("code", "").strip()
    product_name    = request.args.get("product_name", "").strip()
    ingredient      = request.args.get("ingredient", "").strip()
    codes_raw       = request.args.get("codes", "").strip()
    codes           = [c.strip() for c in codes_raw.split(",") if c.strip()]
    force_refresh   = request.args.get("refresh", "0") == "1"
    try:
        current_price = float(request.args.get("current_price", "") or 0) or None
    except ValueError:
        current_price = None

    if not normalized_name:
        return jsonify({"error": "normalized_name 파라미터가 필요합니다."}), 400

    try:
        data = _enrichment_agent.get(
            normalized_name,
            representative_code=code,
            insurance_codes=codes,
            product_name=product_name,
            ingredient=ingredient,
            current_price=current_price,
            force_refresh=force_refresh,
        )
        return jsonify(data)
    except Exception as e:
        logger.error("enrichment 조회 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.post("/api/domestic/enrichment-bulk")
def domestic_enrichment_bulk():
    """POST /api/domestic/enrichment-bulk
    Body: {"items": [{"normalized_name": "...", "product_name": "...", "ingredient": "...",
                      "current_price": 866, "code": "...", "codes": [...]}, ...]}

    선택된 기준약제 + 비교약제들을 한 번에 enrich (캐시 우선).
    허가일·용법·정확한 일일투약비를 비동기로 채우기 위해 사용. 최대 10건.
    """
    from concurrent.futures import ThreadPoolExecutor

    data = request.get_json() or {}
    items = data.get("items") or []
    if not isinstance(items, list):
        return jsonify({"error": "items 는 list 여야 함"}), 400
    items = items[:10]

    def _one(item: dict):
        norm = (item.get("normalized_name") or "").strip()
        if not norm:
            return None, None
        try:
            price = item.get("current_price")
            price_val = float(price) if price is not None else None
            r = _enrichment_agent.get(
                norm,
                representative_code=item.get("code", "") or "",
                insurance_codes=item.get("codes") or [],
                product_name=item.get("product_name", "") or "",
                ingredient=item.get("ingredient", "") or "",
                current_price=price_val,
            )
            return norm, r
        except Exception as e:
            logger.warning("[bulk enrich] %s 실패: %s", norm, e)
            return norm, {"error": str(e)}

    results: dict = {}
    with ThreadPoolExecutor(max_workers=5) as ex:
        for norm, res in ex.map(_one, items):
            if norm is not None:
                results[norm] = res
    return jsonify({"enrichments": results})


@app.get("/api/domestic/change-reason")
def change_reason():
    """
    약가 변동 사유 조회.
    MarketIntelligenceAgent가 PubMed + HIRA/NECA + MA 전문지를 검색하고
    한국 약가 사후관리 4대 기전(적응증 확대/특허 만료/사용량-연동/실거래가 연동)
    프레임으로 GPT-4o가 분석한다. 결과는 JSON 캐시로 저장된다.

    GET /api/domestic/change-reason
        ?drug=키트루다주
        &drug_en=Keytruda
        &date=2022.03.01
        &ingredient=펨브롤리주맙,유전자재조합
        &ingredient_en=pembrolizumab
        &delta_pct=-25.61
        &refresh=0   (1이면 캐시 무시 재분석)
    """
    drug         = request.args.get("drug", "").strip()
    drug_en      = request.args.get("drug_en", "").strip()
    change_date  = request.args.get("date", "").strip()
    ingredient   = request.args.get("ingredient", "").strip()
    ingredient_en = request.args.get("ingredient_en", "").strip()
    insurance_code = request.args.get("insurance_code", "").strip()
    force_refresh = request.args.get("refresh", "0") == "1"

    try:
        delta_pct = float(request.args.get("delta_pct", "0") or "0") or None
    except ValueError:
        delta_pct = None

    if not drug or not change_date:
        return jsonify({"error": "drug, date 파라미터가 필요합니다."}), 400

    # insurance_code 미제공 시 drug 명에서 자동 lookup (LOE 산수 분석에 필요)
    if not insurance_code:
        try:
            matches = db.search_drug(drug, limit=5)
            if matches:
                insurance_code = matches[0].get("insurance_code", "") or ""
        except Exception:
            pass

    result = _mi_agent.analyze_price_change(
        drug_ko=drug,
        drug_en=drug_en or drug,
        ingredient_ko=ingredient,
        ingredient_en=ingredient_en,
        change_date=change_date,
        delta_pct=delta_pct,
        force_refresh=force_refresh,
        insurance_code=insurance_code,
    )

    # ── ReviewAgent 게이트: 결과가 요청·룰에 부합하는지 최종 검증 (최대 1회 재시도) ──
    req_ctx = {"drug": drug, "date": change_date, "delta_pct": delta_pct,
               "ingredient": ingredient}
    verdict = _review_agent.review_price_change_reason(req_ctx, result, MI_RULES_TEXT)
    if not verdict.get("approved", False):
        logger.info("[Review] 1차 거부 — %s", verdict.get("final_verdict", ""))
        # 재시도: 캐시 무시 + corrective_actions 반영
        retry = _mi_agent.analyze_price_change(
            drug_ko=drug, drug_en=drug_en or drug,
            ingredient_ko=ingredient, ingredient_en=ingredient_en,
            change_date=change_date, delta_pct=delta_pct,
            force_refresh=True,
            insurance_code=insurance_code,
        )
        verdict2 = _review_agent.review_price_change_reason(req_ctx, retry, MI_RULES_TEXT)
        if verdict2.get("approved", False):
            result = retry
            verdict = verdict2
        else:
            # 재시도 후 거부 시 분기:
            #  (a) refs ≥ 3 + Tier A 매체 1건 이상 + mechanism 비-unknown
            #      → 분석 자체는 충분 근거 — confidence 만 medium 으로 강등 (기전·reason 유지)
            #  (b) 그 외 — 기존처럼 unknown/low 강제 + reason 추정형 재포장
            result = retry
            refs = result.get("references") or []
            tier_a = sum(1 for r in refs if (r.get("weight") or 0) >= 2.5)
            mech = (result.get("mechanism") or "unknown").lower()
            sufficient = (
                len(refs) >= 3
                and tier_a >= 1
                and mech not in ("unknown", "")
            )
            if sufficient:
                logger.info(
                    "[Review] 재시도 거부했으나 refs=%d(Tier A=%d) · mech=%s — "
                    "confidence=medium 강등하고 분석 유지",
                    len(refs), tier_a, mech,
                )
                result["confidence"] = "medium"
                result["notes"] = (
                    (result.get("notes", "") + " · ReviewAgent 거부 (근거 충분 → confidence 강등) — "
                     + verdict2.get("final_verdict", "")).strip(" ·")
                )
            else:
                logger.info("[Review] 재시도 거부 + 근거 부족 → unknown 하향")
                result["mechanism"] = "unknown"
                result["mechanism_label"] = "미분류"
                result["confidence"] = "low"
                win = result.get("window", {}) or {}
                win_txt = f"{win.get('from','')}~{win.get('to','')}"
                original_reason = (result.get("reason") or "").strip()
                result["reason"] = (
                    f"추정: 변동 시점 윈도우({win_txt}) 내 공개 보도에서 단일 기전을 확정할 수 없음. "
                    f"패널 리뷰어(OpenAI·Gemini)가 근거 부족 또는 윈도우 정합성 불일치로 거부함. "
                    + (f"1차 분석 요지: {original_reason[:160]}…" if original_reason else "")
                ).strip()
                result["notes"] = (
                    (result.get("notes", "") + " · ReviewAgent 거부 — "
                     + verdict2.get("final_verdict", "")).strip(" ·")
                )
            verdict = verdict2

    result["review"] = verdict

    # RSA 자산 — registry 우선 (skill ground-truth) → drug_enrichment fallback.
    # RSA 면 confidence 자동 강등 + reason 본문에 ⚠ 경고 prepend.
    try:
        from agents.kr_rsa_registry import lookup_rsa
        rsa_info = lookup_rsa(drug)
        if not (rsa_info and rsa_info.get("is_rsa") == 1):
            row = _enrichment_agent.db.get_enrichment(drug)
            if row and row.get("is_rsa"):
                rsa_info = {
                    "is_rsa": row.get("is_rsa"),
                    "rsa_type": row.get("rsa_type"),
                    "rsa_note": row.get("rsa_note") or "",
                    "source": "drug_enrichment (Perplexity)",
                }
        if rsa_info and rsa_info.get("is_rsa") == 1:
            existing_conf = (result.get("confidence") or "").lower()
            if existing_conf == "high":
                result["confidence"] = "medium"
            warning = (
                "⚠ 위험분담제(RSA) 자산 — 표시가 ≠ 실제가. "
                "정부와의 RSA 계약 하 표시가는 부분 조정만 노출되며 환급·총액제한 차액은 비공개. "
                "표시가 변동률만으로 mechanism 을 단정할 수 없음.\n\n"
            )
            existing_reason = (result.get("reason") or "").strip()
            if not existing_reason.startswith("⚠"):
                result["reason"] = warning + existing_reason
            result["is_rsa"] = rsa_info.get("is_rsa")
            result["rsa_type"] = rsa_info.get("rsa_type")
            result["rsa_note"] = rsa_info.get("rsa_note") or ""
            result["rsa_source"] = rsa_info.get("source")
    except Exception as e:
        logger.debug("RSA 강등 처리 실패: %s", e)

    return jsonify(result)


@app.get("/api/domestic/media-leaderboard")
def media_leaderboard():
    """매체 신뢰도 리더보드 조회 (캘리브레이션 날짜 포함)."""
    return jsonify(_mi_agent.get_media_leaderboard())


# ──────────────────────────────────────────────────────────────────────────────
# 매체 신뢰도 캘리브레이션 (분기 1회 권장)
# ──────────────────────────────────────────────────────────────────────────────

@app.post("/api/admin/calibrate-media")
def calibrate_media():
    """
    MediaCalibrator 실행 — 10개 기준 약제로 매체 신뢰도 재측정.
    dry_run=true이면 기사 수집만 하고 저장하지 않음.

    POST /api/admin/calibrate-media
    Body: {"dry_run": false}

    주의: 실행에 5~10분 소요됨 (DuckDuckGo 검색 + GPT-4o 평가).
    """
    import threading

    body    = request.get_json(silent=True) or {}
    dry_run = body.get("dry_run", False)

    def _run():
        from agents.media_calibrator import run_calibration
        try:
            result = run_calibration(dry_run=dry_run)
            logger.info("[Calibrator] 완료: %s", result.get("saved_path", "dry-run"))
            # 완료 후 에이전트 가중치 즉시 갱신
            if not dry_run:
                from agents.market_intelligence import _apply_calibrated_weights
                _apply_calibrated_weights()
        except Exception as e:
            logger.error("[Calibrator] 실패: %s", e, exc_info=True)

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()

    return jsonify({
        "status": "started",
        "message": "캘리브레이션이 백그라운드에서 실행 중입니다. "
                   "완료까지 5~10분 소요됩니다.",
        "dry_run": dry_run,
    })


@app.get("/api/admin/rsa-registry")
def rsa_registry_list():
    """RSA registry 전체 조회 (admin 전용 → 추후 auth gate 추가 가능)."""
    from agents.kr_rsa_registry import list_all
    reg = list_all()
    return jsonify({
        "count": len(reg),
        "entries": [
            {"brand_key": k, **v}
            for k, v in sorted(reg.items())
        ],
    })


@app.post("/api/admin/rsa-registry")
def rsa_registry_upsert():
    """RSA registry 추가/수정.

    Body: {"brand_key":"신약X", "is_rsa":1, "rsa_type":"refund", "rsa_note":"...", "source":"user_added"}
    """
    from agents.kr_rsa_registry import add_or_update_rsa
    data = request.get_json(silent=True) or {}
    try:
        entry = add_or_update_rsa(
            brand_key=(data.get("brand_key") or "").strip(),
            is_rsa=int(data.get("is_rsa", 1)),
            rsa_type=data.get("rsa_type") or None,
            rsa_note=data.get("rsa_note") or "",
            source=data.get("source") or "user_added (admin endpoint)",
        )
        return jsonify({"ok": True, "entry": entry})
    except (ValueError, TypeError) as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.delete("/api/admin/rsa-registry/<path:brand_key>")
def rsa_registry_delete(brand_key: str):
    """RSA registry entry 제거."""
    from agents.kr_rsa_registry import remove_rsa
    removed = remove_rsa(brand_key)
    return jsonify({"ok": removed, "brand_key": brand_key})


@app.get("/api/admin/calibration-status")
def calibration_status():
    """가장 최근 캘리브레이션 결과 요약 조회."""
    from agents.media_calibrator import load_latest_calibration
    cal = load_latest_calibration()
    if not cal:
        return jsonify({"status": "미보정", "message": "캘리브레이션 결과 없음"})
    return jsonify({
        "status": "완료",
        "calibrated_at": cal["calibrated_at"],
        "drug_count":    cal["drug_count"],
        "updated_media": len(cal.get("weight_updates", {})),
        "top_media": sorted(
            [
                {"media": name, "new_weight": info["new_weight"],
                 "old_weight": info["old_weight"]}
                for name, info in cal.get("weight_updates", {}).items()
            ],
            key=lambda x: -x["new_weight"]
        )[:5],
    })


@app.get("/dashboard/")
@app.get("/dashboard")
def serve_dashboard_index():
    """메인 대쉬보드 (통합 탭 뷰)."""
    return send_from_directory(str(BASE_DIR / "data" / "dashboard"), "index.html")


@app.get("/dashboard/<path:filename>")
def serve_dashboard(filename: str):
    """대쉬보드 파일 서빙."""
    return send_from_directory(str(BASE_DIR / "data" / "dashboard"), filename)


# ──────────────────────────────────────────────────────────────────────────────
# 해외 약가 검색
# ──────────────────────────────────────────────────────────────────────────────

@app.post("/api/foreign/search")
def foreign_search():
    """
    해외 약가 실시간 검색 (스크레이핑 실행)
    POST /api/foreign/search
    Body: {"query": "Keytruda", "countries": ["JP"], "use_cache": false}

    - use_cache=true: DB에 저장된 이전 결과 반환 (스크레이핑 없음)
    - use_cache=false: 실시간 스크레이핑 후 DB 저장 (기본)
    """
    body = request.get_json(silent=True) or {}
    query = body.get("query", "").strip()
    countries = body.get("countries") or AVAILABLE_COUNTRIES
    use_cache = body.get("use_cache", False)

    if not query:
        return jsonify({"error": "검색어(query)를 입력하세요."}), 400

    # 지원하지 않는 국가 필터링
    unsupported = [c for c in countries if c not in AVAILABLE_COUNTRIES]
    supported = [c for c in countries if c in AVAILABLE_COUNTRIES]

    if use_cache:
        # 캐시된 결과만 반환
        cached = foreign_agent.get_cached_results(query)
        return jsonify({
            "query": query,
            "mode": "cache",
            "results": cached,
            "unsupported_countries": unsupported,
        })

    if not supported:
        return jsonify({
            "error": "현재 구현된 국가가 없습니다.",
            "available": AVAILABLE_COUNTRIES,
            "requested": countries,
        }), 422

    # 실시간 스크레이핑 (async → sync 변환)
    loop = asyncio.new_event_loop()
    try:
        results = loop.run_until_complete(
            foreign_agent.search_all(query, countries=supported)
        )
    except Exception as e:
        logger.error("해외 약가 검색 오류: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500
    finally:
        loop.close()

    # coverage_notes: 국가별 빈 결과의 원인 (정책/로그인/스크래핑 실패 구분)
    coverage_notes = _compute_coverage_notes(query, supported, results)

    return jsonify({
        "query": query,
        "mode": "live",
        "results": results,
        "unsupported_countries": unsupported,
        "coverage_notes": coverage_notes,
    })


def _compute_coverage_notes(query: str, countries: list, results) -> dict:
    """국가별로 결과가 비어있으면 country_coverage 의 정책 메타 반환."""
    from agents.scrapers.country_coverage import lookup_policy
    slug = (query or "").strip().lower()
    notes: dict = {}
    # results 가 list 이면 {country: [...]} 구조가 아닐 수 있어 방어적 처리
    if isinstance(results, dict):
        for cc in countries:
            rows = results.get(cc, []) or []
            has_price = any(
                isinstance(r, dict) and r.get("local_price") is not None for r in rows
            )
            if not has_price:
                policy = lookup_policy(cc, slug)
                if policy:
                    notes[cc] = policy
    return notes


@app.get("/api/foreign/cached")
def foreign_cached():
    """
    해외 약가 결과 조회
    GET /api/foreign/cached?q=Keytruda[&use_cache=false]

    use_cache=false: 실시간 스크래이핑 실행 후 최신 데이터 반환
    use_cache=true (기본): DB 캐시된 데이터만 반환
    """
    query = request.args.get("q", "").strip()
    use_cache = request.args.get("use_cache", "true").lower() != "false"

    if not query:
        return jsonify({"error": "검색어(q)를 입력하세요."}), 400

    # use_cache=false 일 때: 실시간 스크래이핑 실행
    if not use_cache:
        logger.info("[API] 실시간 재검색 시작: %s", query)
        try:
            results = asyncio.run(foreign_agent.search_all(query))
            logger.info("[API] 실시간 검색 완료: %s (%d개 국가)", query, len(results))
        except Exception as e:
            logger.error("[API] 실시간 검색 실패: %s", e, exc_info=True)
            return jsonify({"error": f"실시간 검색 실패: {str(e)}"}), 500

    # DB에서 최신 데이터 조회 (실시간 검색했으면 방금 저장된 데이터)
    cached = foreign_agent.get_cached_results(query)

    # 검색 이력 기록
    try:
        total = sum(len(v) if isinstance(v, list) else 0 for v in cached.values())
        db.log_search(query, "foreign_price", result_count=total)
    except Exception:
        pass

    coverage_notes = _compute_coverage_notes(query, AVAILABLE_COUNTRIES, cached)
    return jsonify({
        "query": query,
        "results": cached,
        "refreshed": not use_cache,
        "coverage_notes": coverage_notes,
    })


@app.get("/api/foreign/drugs")
def foreign_drug_list():
    """
    지금까지 검색된 해외 약제 목록 (검색 히스토리 사이드바용).
    GET /api/foreign/drugs
    반환: [{"query_name", "last_searched_at", "country_count", "has_price"}]
    """
    return jsonify(db.get_foreign_drug_list())


@app.delete("/api/foreign/drugs/<query_name>")
def foreign_drug_delete(query_name: str):
    """검색 이력 삭제. DELETE /api/foreign/drugs/<query_name>
    A8 가격·HTA·허가 모든 캐시를 함께 지운다. 복구 불가.
    """
    from urllib.parse import unquote
    qn = unquote(query_name).strip()
    if not qn:
        return jsonify({"error": "empty query_name"}), 400
    deleted = db.delete_foreign_drug(qn)
    # HTA / 허가 캐시도 alias 전부 정리 (있을 때만)
    try:
        from agents.db.drug_aliases import aliases as _aliases
        names = _aliases(qn)
        placeholders = ",".join(["?"] * len(names))
        with db._connect() as conn:
            conn.execute(
                f"DELETE FROM hta_approvals_cache WHERE LOWER(drug_query) IN ({placeholders})",
                tuple(names),
            )
    except Exception:
        pass
    return jsonify({"ok": True, "deleted": deleted, "query_name": qn})


@app.get("/api/foreign/available_countries")
def available_countries():
    return jsonify({"available": AVAILABLE_COUNTRIES})


# ──────────────────────────────────────────────────────────────────────────────
# /api/foreign/country-overview — 국가별 카드 그리드용 통합 응답
# (pure-napping-goose plan / Phase 4)
# axis: brand 검색 → 국가 카드 (각 카드: 허가 N건 + 급여 status + 가격)
# ──────────────────────────────────────────────────────────────────────────────

# 허가 agency ↔ 급여 body 매핑 (1국 1 agency 가정. EU 는 EMA 만 있음)
_COUNTRY_AGENCY_MAP = [
    # (country, agency_for_approval, body_for_reimbursement, currency_hint)
    ("US", "FDA",  "CMS",     "USD"),
    ("EU", "EMA",  None,      "EUR"),
    ("UK", "MHRA", "NICE",    "GBP"),
    ("JP", "PMDA", "CHUIKYO", "JPY"),
    ("AU", "TGA",  "PBAC",    "AUD"),
    ("KR", "MFDS", "HIRA",    "KRW"),
]


def _summarize_reimbursement(rows: list[dict]) -> str:
    """국가 카드 헤더 pill 라벨. 우선순위: recommend > restrict > optimised > reject > not_listed > not_applicable > none."""
    if not rows:
        return "none"
    types = {r.get("decision_type") for r in rows if r.get("decision_type")}
    for level in ("recommend", "restrict", "optimised", "reject", "not_listed", "not_applicable"):
        if level in types:
            return level
    return "none"


def _country_card_row(product_slug: str, country: str, agency: str | None,
                       body: str | None, currency: str | None) -> dict:
    """단일 (product, country) 카드 row 생성.

    indication × agency × reimbursement × price 4-way 조회 결과를 카드용 dict 로.
    """
    # 1) 허가 — indications_by_agency
    with db._connect() as conn:
        approvals = []
        if agency:
            approvals = [dict(r) for r in conn.execute("""
                SELECT m.indication_id, m.title, m.disease, m.line_of_therapy,
                       m.biomarker_class, ia.approval_date, ia.label_excerpt,
                       ia.label_url, ia.restriction_note
                  FROM indications_master m
                  JOIN indications_by_agency ia ON m.indication_id = ia.indication_id
                 WHERE m.product = ? AND ia.agency = ?
                 ORDER BY ia.approval_date DESC NULLS LAST
            """, (product_slug, agency)).fetchall()]

    # 2) 급여 — reimbursement_xnational + indication_reimbursement (KR)
    if country == "KR":
        # KR: indication_reimbursement union 으로 처리
        reimb_rows = db.get_xnational_reimbursement_for_product(product_slug)
        reimb_rows = [r for r in reimb_rows if r.get("country") == "KR"]
    else:
        reimb_rows = db.get_xnational_reimbursement_for_product(product_slug)
        reimb_rows = [r for r in reimb_rows if r.get("country") == country]

    # 3) 가격 — foreign_drug_prices (alias_map 기반 조회). KR 은 별도 (drug_prices)
    price_rows: list[dict] = []
    if country != "KR":
        # alias 기반 검색 — get_foreign_prices 가 그대로 작동
        try:
            all_prices = db.get_foreign_prices(product_slug)
            price_rows = [p for p in all_prices if p.get("country") == country]
        except Exception:
            price_rows = []

    # 4) indications array — 허가 row 에 같은 indication_id 의 reimbursement / price 결합
    by_ind: dict[str, dict] = {}
    for a in approvals:
        ind_id = a["indication_id"]
        by_ind[ind_id] = {
            "indication_id": ind_id,
            "title":         a.get("title"),
            "disease":       a.get("disease"),
            "line_of_therapy": a.get("line_of_therapy"),
            "biomarker":     a.get("biomarker_class"),
            "approval_date": a.get("approval_date"),
            "label_excerpt": (a.get("label_excerpt") or "")[:300],
            "label_url":     a.get("label_url"),
            "reimbursement": None,
            "price":         None,
        }
    # reimbursement attach
    for r in reimb_rows:
        ind_id = r.get("indication_id")
        if ind_id and ind_id in by_ind:
            by_ind[ind_id]["reimbursement"] = {
                "decision_type": r.get("decision_type"),
                "decision_id":   r.get("decision_id"),
                "decision_date": r.get("decision_date"),
                "criteria_text": (r.get("criteria_text") or "")[:500],
                "source_url":    r.get("source_url"),
                "body":          r.get("body"),
            }
    # price attach — 가격은 indication-specific 이 아닐 수도 (brand-level). 첫 가격을 모든 indication 에 attach
    rep_price = None
    if price_rows:
        # 가장 최신 / form_type 우선순위: oral > injection > unknown
        def _form_priority(p):
            ft = (p.get("form_type") or "").lower()
            return {"oral": 0, "injection": 1}.get(ft, 2)
        pr = sorted(price_rows, key=lambda p: (_form_priority(p), p.get("searched_at") or ""), reverse=False)[0]
        rep_price = {
            "currency":            pr.get("currency"),
            "local_price":         pr.get("local_price"),
            "adjusted_price_krw":  pr.get("adjusted_price_krw"),
            "daily_cost_krw":      pr.get("daily_cost_krw"),
            "form_type":           pr.get("form_type"),
            "dosage_strength":     pr.get("dosage_strength"),
            "package_unit":        pr.get("package_unit"),
            "source_label":        pr.get("source_label"),
            "searched_at":         pr.get("searched_at"),
        }
        for v in by_ind.values():
            v["price"] = rep_price

    # KR 가격은 별도 — drug_prices.search_drug 에서 brand 기준으로 첫 row
    if country == "KR" and not rep_price:
        try:
            with db._connect() as conn:
                rk = conn.execute("""
                    SELECT max_price, dosage_strength, package_unit
                      FROM drug_latest
                     WHERE LOWER(product_name_kr) LIKE ?
                     LIMIT 1
                """, (f"%{product_slug}%",)).fetchone()
            if rk:
                rep_price = {
                    "currency": "KRW",
                    "local_price": rk[0],
                    "adjusted_price_krw": rk[0],
                    "daily_cost_krw": None,
                    "form_type": None,
                    "dosage_strength": rk[1],
                    "package_unit": rk[2],
                    "source_label": "HIRA",
                    "searched_at": None,
                }
                for v in by_ind.values():
                    v["price"] = rep_price
        except Exception:
            pass

    indications_list = sorted(
        by_ind.values(),
        key=lambda v: (v.get("approval_date") or "0000-00-00"),
        reverse=True,
    )

    return {
        "country": country,
        "agency":  agency,
        "body":    body,
        "currency_hint": currency,
        "approval_count": len(approvals),
        "indications":    indications_list,
        "reimbursement_summary": _summarize_reimbursement(reimb_rows),
        "reimbursement_count":   len(reimb_rows),
        "price_summary":         rep_price,
    }


@app.get("/api/foreign/country-overview")
def foreign_country_overview():
    """
    국가별 카드 그리드용 단일 통합 응답.
    GET /api/foreign/country-overview?query=keytruda
    """
    query = request.args.get("query", "").strip()
    if not query:
        return jsonify({"error": "query parameter required"}), 400

    # query → product_slug (alias_map 우선, fallback canonical)
    product_slug = query.lower()
    alias = db.get_product_alias(product_slug)
    if not alias:
        # alias_map 의 brand_aliases 안에 query 가 있을 수 있음 — 전체 스캔
        for entry in db.list_product_aliases():
            keys = [entry["product_slug"].lower()]
            if entry.get("inn"):
                keys.append(entry["inn"].lower())
            keys += [str(x).lower() for x in entry.get("brand_aliases", [])]
            if query.lower() in keys:
                product_slug = entry["product_slug"]
                alias = entry
                break

    inn = (alias or {}).get("inn")

    countries = []
    for country, agency, body, currency in _COUNTRY_AGENCY_MAP:
        countries.append(_country_card_row(product_slug, country, agency, body, currency))

    return jsonify({
        "product": product_slug,
        "inn":     inn,
        "query":   query,
        "countries": countries,
    })


@app.get("/api/search/history")
def search_history():
    """검색 이력 조회. GET /api/search/history?type=hta&limit=20"""
    search_type = request.args.get("type")
    limit = int(request.args.get("limit", 20))
    return jsonify(db.get_search_history(search_type, limit))


@app.get("/api/data/freshness")
def data_freshness():
    """데이터 신선도 조회. GET /api/data/freshness?type=hta&key=belzutifan_FDA"""
    data_type = request.args.get("type", "").strip()
    scope_key = request.args.get("key", "").strip()
    if data_type and scope_key:
        info = db.get_freshness(data_type, scope_key)
        return jsonify(info or {"status": "not_found"})
    # 전체 목록
    with db._connect() as conn:
        rows = conn.execute("SELECT * FROM data_freshness ORDER BY last_fetched DESC").fetchall()
    return jsonify([dict(r) for r in rows])


# ──────────────────────────────────────────────────────────────────────────────
# HIRA 에이전트 (2025.3월 개정 SOP + 외국약가 조정가 검증)
# ──────────────────────────────────────────────────────────────────────────────

from agents.hira_agent import HIRAAgent
_hira_agent = HIRAAgent()


@app.get("/api/hira/pricing-summary")
def hira_pricing_summary():
    """약제결정신청(요양급여 등재) 핵심 조항 요약."""
    try:
        return jsonify(_hira_agent.pricing_application_summary())
    except Exception as e:
        logger.error("HIRA summary 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/hira/checklist")
def hira_checklist():
    return jsonify({"items": _hira_agent.submission_checklist()})


@app.get("/api/hira/article")
def hira_article():
    """GET /api/hira/article?label=제3조의2"""
    label = request.args.get("label", "").strip()
    if not label:
        return jsonify({"error": "label 파라미터 필요 (예: 제3조의2)"}), 400
    art = _hira_agent.get_article(label)
    if not art:
        return jsonify({"error": f"조항 없음: {label}"}), 404
    return jsonify({"label": art.label, "title": art.title, "page": art.page, "body": art.body})


@app.get("/api/hira/audit-adjustment")
def hira_audit_adjustment():
    """_resource/산출식.xlsx 의 수식·비율이 규정과 일치하는지 더블체크."""
    return jsonify(_hira_agent.audit_adjustment_excel())


@app.post("/api/hira/compute-a8")
def hira_compute_a8():
    """
    외국약가 → A8 조정가 산출.
    Body: {
      "prices": {"UK": 132.63, "US": 339.46, ...},   // 최소단위당 현지 통화
      "fx_rates": {"UK": 1821.01, ...},              // optional, 기본값 2025.3월 기준
      "subset": ["UK","US","CA","JP","FR","DE","IT","CH"]  // optional, 최저가 산출 대상
    }
    """
    body = request.get_json(silent=True) or {}
    prices = body.get("prices") or {}
    if not prices:
        return jsonify({"error": "prices 는 필수 — {국가코드: 현지가격} 형식"}), 400
    try:
        result = _hira_agent.compute_a8(
            prices_local=prices,
            fx_rates=body.get("fx_rates"),
            subset=body.get("subset"),
        )
        return jsonify(result)
    except Exception as e:
        logger.error("A8 산출 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# 약제명 리졸버 (제품명 ↔ 성분명)
# ──────────────────────────────────────────────────────────────────────────────

# 국내 DB 에 없는 해외 신약용 최소 매핑 (INN ↔ 대표 상품명).
# 확장 필요 시 data/resource/drug_alias.json 으로 분리 가능.
_KNOWN_ALIASES = {
    "belzutifan":    ["welireg"],
    "pembrolizumab": ["keytruda", "키트루다"],
    "nivolumab":     ["opdivo", "옵디보"],
    "atezolizumab":  ["tecentriq", "티쎈트릭"],
    "durvalumab":    ["imfinzi", "임핀지"],
    "lenvatinib":    ["lenvima", "렌비마"],
    "osimertinib":   ["tagrisso", "타그리소"],
    "sotorasib":     ["lumakras", "lumykras"],
    "enfortumab vedotin": ["padcev", "패드세브"],
}
_ALIAS_TO_INGREDIENT = {}
for ing, products in _KNOWN_ALIASES.items():
    for p in products + [ing]:
        _ALIAS_TO_INGREDIENT[p.lower()] = ing


def _resolve_drug(query: str) -> dict:
    """입력(제품명·성분명·한글·영문)을 성분명 + 제품명 리스트로 해석."""
    q = (query or "").strip()
    if not q:
        return {"query": q, "ingredient": "", "products": [], "source": "empty"}

    # 1) 국내 DB drug_latest 에서 ingredient 검색
    try:
        import sqlite3 as _sqlite3
        like = f"%{q}%"
        with _sqlite3.connect(str(db.db_path)) as c:
            cur = c.execute(
                "SELECT DISTINCT ingredient, product_name_kr, product_name_en "
                "FROM drug_latest "
                "WHERE ingredient LIKE ? OR product_name_kr LIKE ? OR product_name_en LIKE ? "
                "LIMIT 20",
                (like, like, like),
            )
            rows = cur.fetchall()
        if rows:
            ingredients = [r[0] for r in rows if r[0]]
            products = [p for r in rows for p in (r[1], r[2]) if p]
            from collections import Counter
            top_ing = Counter(ingredients).most_common(1)[0][0] if ingredients else ""
            # DB에서 성분명을 못 찾았으면 alias 맵에서 product→ingredient 역매핑 시도
            if not top_ing:
                for prod in products + [q]:
                    ing = _ALIAS_TO_INGREDIENT.get(str(prod).lower().split("(")[0].strip())
                    if ing:
                        top_ing = ing
                        break
            return {
                "query": q,
                "ingredient": top_ing,
                "products": list(dict.fromkeys(products))[:10],
                "source": "domestic_db" + ("+alias" if top_ing and not ingredients else ""),
            }
    except Exception as e:
        logger.warning("resolve domestic lookup 실패: %s", e)

    # 2) 알려진 INN↔상품 매핑
    ing = _ALIAS_TO_INGREDIENT.get(q.lower())
    if ing:
        return {
            "query": q,
            "ingredient": ing,
            "products": _KNOWN_ALIASES.get(ing, []),
            "source": "alias_map",
        }

    # 3) fallback — 입력을 성분명으로 간주
    return {"query": q, "ingredient": q, "products": [], "source": "fallback"}


@app.get("/api/drug/resolve")
def drug_resolve():
    q = (request.args.get("q") or "").strip()
    if not q:
        return jsonify({"error": "q 는 필수"}), 400
    return jsonify(_resolve_drug(q))


# ──────────────────────────────────────────────────────────────────────────────
# HTA 허가현황 (PBAC/CADTH/NICE/SMC)
# ──────────────────────────────────────────────────────────────────────────────

from agents.hta_approval_agent import HTAApprovalAgent
_hta_agent = HTAApprovalAgent()


@app.get("/api/hta/approvals")
def hta_approvals():
    """GET /api/hta/approvals?drug=belzutifan[&body=SMC][&refresh=1]"""
    drug = (request.args.get("drug") or "").strip()
    if not drug:
        return jsonify({"error": "drug 는 필수"}), 400
    body = (request.args.get("body") or "").strip().upper() or None
    refresh = request.args.get("refresh") in ("1", "true", "True")
    try:
        results = _hta_agent.get(drug, body=body, force_refresh=refresh)
        return jsonify({
            "drug": drug,
            "available_bodies": _hta_agent.available_bodies(),
            "count": len(results),
            "results": results,
        })
    except Exception as e:
        logger.error("HTA 조회 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/hta/pdf")
def hta_pdf():
    """GET /api/hta/pdf?path=<pdf_local 절대경로>  — 캐시된 평가 PDF 다운로드"""
    from flask import send_file, abort
    p = request.args.get("path") or ""
    if not p:
        abort(400)
    fp = Path(p).resolve()
    cache_root = (Path(__file__).parent.parent / "data" / "hta_cache").resolve()
    if not str(fp).startswith(str(cache_root)) or not fp.exists():
        abort(404)
    return send_file(str(fp), mimetype="application/pdf", as_attachment=False, download_name=fp.name)


@app.get("/api/hta/indication-matrix")
def hta_indication_matrix():
    """GET /api/hta/indication-matrix?drug=belzutifan[&refresh=1]

    FDA 적응증을 축으로 PBAC/CADTH/NICE/SMC 평가를 매칭한 매트릭스 반환.
    캐시 데이터가 있으면 즉시 반환 (refresh=1 시에만 재수집).
    """
    drug = (request.args.get("drug") or "").strip()
    if not drug:
        return jsonify({"error": "drug 는 필수"}), 400
    refresh = request.args.get("refresh") in ("1", "true", "True")
    try:
        matrix = _hta_agent.get_indication_matrix(drug, force_refresh=refresh)
        # 검색 이력 + 신선도 기록
        try:
            n_ind = len(matrix.get("indications", []))
            db.log_search(drug, "hta", result_count=n_ind)
            if n_ind > 0:
                for body in ["FDA", "PBAC", "CADTH", "NICE", "SMC"]:
                    db.update_freshness("hta", f"{drug}_{body}")
        except Exception:
            pass
        return jsonify(matrix)
    except Exception as e:
        logger.error("Indication matrix 조회 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# Regulatory Approval Matrix (6-agency)
# ──────────────────────────────────────────────────────────────────────────────

from agents.foreign_approval import ForeignApprovalAgent
_approval_agent = ForeignApprovalAgent()

@app.get("/api/approval/matrix")
def approval_matrix():
    """GET /api/approval/matrix?product=keytruda"""
    product = (request.args.get("product") or "").strip().lower()
    if not product:
        return jsonify({"error": "product 는 필수"}), 400
    try:
        m = _approval_agent.matrix(product)
        return jsonify(m)
    except Exception as e:
        logger.error("Approval matrix 조회 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


_TRANSLATE_CACHE_DIR = BASE_DIR / "data" / "hta_cache" / "translations"
_TRANSLATE_CACHE_DIR.mkdir(parents=True, exist_ok=True)


def _translate_ja_to_ko(text: str) -> str:
    """Gemini 2.5-flash 로 일본어 → 한국어 번역 (파일 캐시)."""
    if not text or not text.strip():
        return text
    import hashlib, json as _json, urllib.request, ssl, os
    cache_key = hashlib.md5(text.encode()).hexdigest()
    cache_file = _TRANSLATE_CACHE_DIR / f"ja_ko_{cache_key}.txt"
    if cache_file.exists():
        return cache_file.read_text("utf-8")

    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        return text

    body = {
        "contents": [{"parts": [{"text":
            f"다음 일본어 의약품 허가사항을 한국어로 번역하세요. 의약품/질환 전문용어를 정확히 사용하세요. 번역문만 출력하세요.\n\n{text}"
        }]}],
        "generationConfig": {"temperature": 0, "maxOutputTokens": 1024},
    }
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    try:
        req = urllib.request.Request(url, data=_json.dumps(body).encode("utf-8"),
                                     headers={"Content-Type": "application/json"}, method="POST")
        ctx = ssl.create_default_context()
        with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
            payload = _json.loads(resp.read().decode("utf-8"))
        translated = (payload.get("candidates", [{}])[0]
                      .get("content", {}).get("parts", [{}])[0]
                      .get("text", "")).strip()
        if translated:
            cache_file.write_text(translated, "utf-8")
            return translated
    except Exception as e:
        logger.warning("번역 실패: %s", e)
    return text


def _is_japanese(text: str) -> bool:
    if not text:
        return False
    for ch in text[:100]:
        if '\u3040' <= ch <= '\u30ff' or '\u4e00' <= ch <= '\u9fff':
            return True
    return False


@app.get("/api/approval/full_text")
def approval_full_text():
    """GET /api/approval/full_text?product=keytruda

    product slug 의 모든 적응증 × agency 조합에 대한 허가 원문 전문 반환.
    대쉬보드 '허가 문구' 패널이 요약 대신 원문을 보여주기 위해 사용.
    """
    product = (request.args.get("product") or "").strip().lower()
    if not product:
        return jsonify({"error": "product 는 필수"}), 400
    try:
        rows: list[dict] = []
        with db._connect() as conn:
            for r in conn.execute(
                "SELECT m.indication_id, m.disease, m.stage, m.line_of_therapy, "
                "       m.biomarker_class, m.title, "
                "       a.agency, a.approval_date, a.date_source, a.label_excerpt, a.label_full_text, "
                "       a.label_url, a.biomarker_label, a.combination_label "
                "FROM indications_master m "
                "JOIN indications_by_agency a ON a.indication_id = m.indication_id "
                "WHERE m.product = ? "
                "ORDER BY (a.approval_date IS NULL), a.approval_date DESC, a.agency, m.disease",
                (product,),
            ):
                d = dict(r)
                if d.get("agency") == "PMDA":
                    ft = d.get("label_full_text") or ""
                    if _is_japanese(ft):
                        d["label_full_text_original"] = ft
                        d["label_full_text"] = _translate_ja_to_ko(ft)
                    ex = d.get("label_excerpt") or ""
                    if _is_japanese(ex):
                        d["label_excerpt_original"] = ex
                        d["label_excerpt"] = _translate_ja_to_ko(ex)
                rows.append(d)
        by_agency: dict[str, list[dict]] = {}
        for row in rows:
            by_agency.setdefault(row["agency"], []).append(row)
        return jsonify({"product": product, "by_agency": by_agency})
    except Exception as e:
        logger.error("Approval full_text 조회 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/approval/detail")
def approval_detail():
    """GET /api/approval/detail?id=keytruda_nsclc_1l_metastatic_chemo"""
    ind_id = (request.args.get("id") or "").strip()
    if not ind_id:
        return jsonify({"error": "id 는 필수"}), 400
    try:
        rec = db.get_indication(ind_id)
        if not rec:
            return jsonify({"error": "not found"}), 404
        product = rec.get("product")
        initial_auth: dict[str, str] = {}
        if product:
            with db._connect() as conn:
                for row in conn.execute(
                    "SELECT a.agency, MIN(a.approval_date) "
                    "FROM indications_by_agency a "
                    "JOIN indications_master m ON m.indication_id = a.indication_id "
                    "WHERE m.product = ? AND a.approval_date IS NOT NULL "
                    "GROUP BY a.agency",
                    (product,),
                ):
                    initial_auth[row[0]] = row[1]
        for a in (rec.get("agencies") or []):
            ag = a.get("agency")
            a["initial_auth_date"] = initial_auth.get(ag)
            if ag == "PMDA":
                excerpt = a.get("label_excerpt") or ""
                if _is_japanese(excerpt):
                    a["label_excerpt_original"] = excerpt
                    a["label_excerpt"] = _translate_ja_to_ko(excerpt)
                full_text = a.get("label_full_text") or ""
                if _is_japanese(full_text):
                    a["label_full_text_original"] = full_text
                    a["label_full_text"] = _translate_ja_to_ko(full_text)
                combo = a.get("combination_label") or ""
                if _is_japanese(combo):
                    a["combination_label_original"] = combo
                    a["combination_label"] = _translate_ja_to_ko(combo)
        return jsonify(rec)
    except Exception as e:
        logger.error("Approval detail 조회 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/approval/products")
def approval_products():
    """등록된 product 목록 + 간단한 통계."""
    try:
        with db._connect() as conn:
            rows = conn.execute(
                "SELECT DISTINCT product FROM indications_master ORDER BY product"
            ).fetchall()
        products = []
        for r in rows:
            slug = r[0]
            m = _approval_agent.matrix(slug)
            products.append({
                "product": slug,
                "masters": m["totals"]["masters"],
                "agencies": {
                    "FDA": m["totals"]["fda_agency"],
                    "EMA": m["totals"]["ema_agency"],
                    "PMDA": m["totals"]["pmda_agency"],
                    "MFDS": m["totals"]["mfds_agency"],
                    "MHRA": m["totals"]["mhra_agency"],
                    "TGA": m["totals"]["tga_agency"],
                },
                "all_six": m["totals"]["all_six"],
            })
        products.sort(key=lambda p: p["masters"], reverse=True)
        return jsonify({"products": products})
    except Exception as e:
        logger.error("Approval products 조회 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# Negotiation Workbench (Phase 1 MVP)
# ──────────────────────────────────────────────────────────────────────────────

from datetime import datetime
from flask import send_file
from agents.workbench import (
    DEFAULT_ASSUMPTIONS,
    compute_all_scenarios,
    export_workbook,
    list_available_products,
    load_assumptions,
    load_hta_for_product,
    save_assumptions,
    summarize_hta,
)


@app.get("/api/workbench/assumptions")
def workbench_assumptions_get():
    """현재 가정치 (없으면 DEFAULT) 반환. 설정 화면용."""
    return jsonify(load_assumptions())


@app.put("/api/workbench/assumptions")
def workbench_assumptions_put():
    """가정치 전체 저장. Body: assumptions dict 전체."""
    body = request.get_json(silent=True) or {}
    if not body or "countries" not in body:
        return jsonify({"error": "countries 키 필수"}), 400
    save_assumptions(body, user=body.get("_user", "dashboard"))
    return jsonify({"ok": True, "saved": load_assumptions()})


@app.get("/api/workbench/defaults")
def workbench_defaults():
    """HIRA 고시 기본값 (복원용)."""
    return jsonify(DEFAULT_ASSUMPTIONS)


@app.get("/api/workbench/hta")
def workbench_hta():
    """
    제품별 Tier-3 HTA 교차검증 캐시 조회.
    Query: ?product=keytruda&summary=1 (summary=1 이면 요약만 반환)

    Returns:
      - full=True:  {"data": {nice:..., pbac:..., has:..., gba:...}, "summary": {...}}
      - full=False: {"summary": {...}}
    """
    product = request.args.get("product", "").strip()
    only_summary = request.args.get("summary") in ("1", "true")

    if not product:
        # 제품 인자 없으면 사용 가능한 목록 반환
        return jsonify({
            "available_products": list_available_products(),
            "hint": "?product=keytruda",
        })

    data = load_hta_for_product(product)
    if data is None:
        return jsonify({
            "error": f"제품 '{product}' 의 HTA 캐시 없음",
            "available_products": list_available_products(),
        }), 404

    summary = summarize_hta(data)
    if only_summary:
        return jsonify({"product": product, "summary": summary})
    return jsonify({"product": product, "data": data, "summary": summary})


@app.post("/api/workbench/compute")
def workbench_compute():
    """
    시나리오 병렬 계산 + (옵션) dose 정규화.

    Body: {
      "prices":       {"JP": 88300, "IT": 1200, ...},   # 국가별 현지가격
      "rows_meta":    {country: {product_name, strength, pack, form}}  (선택)
      "product_slug": "keytruda"                         (선택, REFERENCE_SKU 폴백용)
      "reference_mg": 100                                (선택, 기준 mg)
      "scenarios":    [...],
      "assumptions":  {...}                              (선택)
    }

    rows_meta 가 있으면 국가별 SKU 의 strength/pack 을 파싱해 equivalent_price
    (reference_mg 기준 환산가) 를 계산 후 A8 비교. 없으면 raw local_price 비교.
    응답의 각 시나리오 rows[country] 에 mg_pack_total/price_per_mg/dose_confidence 추가,
    excluded dict 에 동등비교 불가 국가와 사유 표기.
    """
    body = request.get_json(silent=True) or {}
    prices = body.get("prices") or {}
    if not prices:
        return jsonify({"error": "prices 필수"}), 400
    scenarios = body.get("scenarios") or []
    if not scenarios:
        return jsonify({"error": "scenarios 필수 (최소 1개)"}), 400
    assumptions = body.get("assumptions") or load_assumptions()
    rows_meta    = body.get("rows_meta")
    product_slug = body.get("product_slug")
    reference_mg = body.get("reference_mg")

    try:
        results = compute_all_scenarios(
            prices, scenarios, assumptions,
            rows_meta=rows_meta,
            product_slug=product_slug,
            reference_mg=reference_mg,
        )
        # HTA 캐시 자동 attach (있으면 summary, 없으면 null)
        hta_summary = None
        if product_slug:
            hta_data = load_hta_for_product(product_slug)
            if hta_data:
                hta_summary = summarize_hta(hta_data)
        return jsonify({"scenarios": results, "hta_summary": hta_summary})
    except Exception as e:
        logger.error("workbench compute 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.post("/api/workbench/export")
def workbench_export():
    """
    세션 전체 → xlsx 생성 후 다운로드.
    Body: {
      "project":   {...},
      "prices":    {...},
      "scenarios": [...],   # compute 결과 그대로
      "selected":  "B안",
      "source_raw": [...],
      "matching":  [...],
      "hta":       [...] | null,
      "audit_log": [...],
    }
    """
    body = request.get_json(silent=True) or {}
    if not body.get("scenarios"):
        return jsonify({"error": "scenarios 필수"}), 400

    session = dict(body)
    session.setdefault("assumptions", load_assumptions())

    # 파일명
    proj = session.get("project", {})
    drug = (proj.get("drug_name_en") or proj.get("drug_name_kr") or "product").replace(" ", "_").replace("(", "").replace(")", "")

    # HTA 자동 로드 — 클라이언트가 보내지 않았거나 dict 가 아닌 경우 캐시에서 attach
    if not isinstance(session.get("hta"), dict):
        for key in ("drug_name_en", "drug_name_kr"):
            cached = load_hta_for_product(proj.get(key, ""))
            if cached:
                session["hta"] = cached
                break
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = BASE_DIR / "data" / "workbench" / "exports" / f"MA_A8_Workbench_{drug}_{stamp}.xlsx"

    try:
        export_workbook(session, out_path)
        return send_file(
            str(out_path),
            as_attachment=True,
            download_name=out_path.name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        logger.error("workbench export 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# 한국MSD 요약 (Home 카드용)
# ──────────────────────────────────────────────────────────────────────────────

DISEASE_KR = {
    "NSCLC": "비소세포폐암", "MEL": "흑색종", "HNSCC": "두경부암",
    "cHL": "호지킨림프종", "UC": "요로상피암", "GC": "위암/위식도접합부암",
    "ESC": "식도암", "EC": "자궁내막암", "CC": "자궁경부암", "TNBC": "삼중음성유방암",
    "RCC": "신세포암", "HCC": "간세포암", "CRC": "대장암", "MCC": "메르켈세포암",
    "BTC": "담도암", "MPM": "악성흉막중피종", "PMBCL": "원발성종격동B세포림프종",
    "cSCC": "피부편평세포암", "OC": "난소암", "SOLID": "고형암(MSI-H/TMB-H)",
}

LINE_KR = {
    "1L": "1차", "2L": "2차", "3L+": "3차 이상", "3L": "3차",
    "adjuvant": "보조요법", "neoadjuvant": "신보조요법",
    "perioperative": "수술 전후 요법", "maintenance": "유지요법",
}


@app.get("/api/msd/summary")
def msd_summary():
    """Home 카드용 — MSD 급여 품목 수 (최신 고시일 기준) + Keytruda 적응증 현황."""
    try:
        with db._connect() as conn:
            # 최신 고시일 (apply_date) — drug_prices 기준
            latest_row = conn.execute(
                "SELECT MAX(apply_date) FROM drug_prices"
            ).fetchone()
            latest_apply_date = latest_row[0] if latest_row and latest_row[0] else None

            if latest_apply_date:
                row = conn.execute(
                    "SELECT COUNT(DISTINCT product_name_kr) FROM drug_prices "
                    "WHERE company LIKE ? AND apply_date = ? AND max_price > 0",
                    ("%엠에스디%", latest_apply_date),
                ).fetchone()
            else:
                row = conn.execute(
                    "SELECT COUNT(DISTINCT product_name_kr) FROM drug_latest "
                    "WHERE company LIKE ?",
                    ("%엠에스디%",),
                ).fetchone()
            reimbursed_product_count = int(row[0]) if row else 0

            ind_rows = conn.execute(
                """
                SELECT m.indication_id, m.disease, m.line_of_therapy, m.stage,
                       m.biomarker_class,
                       (SELECT approval_date FROM indications_by_agency
                         WHERE indication_id = m.indication_id AND agency='MFDS') AS mfds_date,
                       (SELECT approval_date FROM indications_by_agency
                         WHERE indication_id = m.indication_id AND agency='FDA') AS fda_date,
                       r.is_reimbursed, r.effective_date, r.criteria_text,
                       r.notice_date, r.notice_url,
                       m.pivotal_trial
                FROM indications_master m
                LEFT JOIN indication_reimbursement r
                    ON r.indication_id = m.indication_id
                WHERE m.product = 'keytruda'
                ORDER BY COALESCE(mfds_date, fda_date, '0000') DESC
                """
            ).fetchall()

        total = len(ind_rows)
        mfds_approved = sum(1 for r in ind_rows if r[5])
        reimbursed_count = sum(1 for r in ind_rows if r[7])
        items = []
        for r in ind_rows:
            disease = r[1] or ""
            lot = r[2] or ""
            stage = r[3] or ""
            bio = r[4] or ""
            disease_kr = DISEASE_KR.get(disease, disease)
            lot_kr = LINE_KR.get(lot, lot) if lot else ""
            parts = [disease_kr]
            if lot_kr:
                parts.append(lot_kr)
            if stage and stage not in ("metastatic", "advanced"):
                parts.append(stage)
            if bio and bio != "all_comers":
                parts.append(bio.replace("_", " "))
            title = " · ".join(parts)
            items.append({
                "id": r[0],
                "disease": disease,
                "disease_kr": disease_kr,
                "line_of_therapy": lot,
                "stage": stage,
                "biomarker_class": bio,
                "title": title,
                "pivotal_trial": r[12] if len(r) > 12 else None,
                "mfds_approved": bool(r[5]),
                "mfds_date": r[5],
                "fda_date": r[6],
                "is_reimbursed": bool(r[7]) if r[7] is not None else False,
                "reimbursement_effective_date": r[8],
                "reimbursement_criteria": r[9],
                "reimbursement_notice_date": r[10],
                "reimbursement_notice_url": r[11],
            })

        return jsonify({
            "reimbursed_product_count": reimbursed_product_count,
            "latest_apply_date": latest_apply_date,
            "keytruda": {
                "total_indications": total,
                "mfds_approved": mfds_approved,
                "pending_mfds": total - mfds_approved,
                "reimbursed_indications": reimbursed_count,
                "pending_reimbursement": total - reimbursed_count,
                "items": items,
            },
        })
    except Exception as e:
        logger.error("MSD summary 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/home/media-intelligence")
def home_media_intelligence():
    """Home 미디어 인텔리전스 카드 — 1개월 브랜드별 트래픽 + 최신뉴스.
    `refresh=1` 쿼리로 캐시 무시 강제 재수집.
    """
    try:
        days_param = request.args.get("days")
        days = int(days_param) if days_param else None  # None → 달력 기반 1개월
        refresh = request.args.get("refresh", "0") == "1"
        data = _media_intel.get_brand_traffic(days=days, refresh=refresh)
        return jsonify(data)
    except Exception as e:
        logger.error("media-intelligence 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/home/brand-news")
def home_brand_news():
    """브랜드 클릭 시 — 오늘 기준 최신 뉴스."""
    brand = request.args.get("brand", "").strip()
    if not brand:
        return jsonify({"error": "brand 파라미터 필요"}), 400
    try:
        limit = int(request.args.get("limit", "10"))
        items = _media_intel.get_latest_brand_news(brand, limit=limit)
        return jsonify({"brand": brand, "count": len(items), "items": items})
    except Exception as e:
        logger.error("brand-news 실패 (%s): %s", brand, e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/home/government-keyword-summary")
def home_government_summary():
    """정부 기관 키워드 (보건복지부/건보공단/심평원/식약처) 최근 1개월 AI 요약.
    OpenAI + Gemini 독립 호출 → 결과 병합. 일자별 cache.
    """
    try:
        from agents.government_keyword_summary import get_government_summary
        refresh = request.args.get("refresh", "0") == "1"
        data = get_government_summary(refresh=refresh)
        return jsonify(data)
    except Exception as e:
        logger.error("government summary 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/home/top-price-changes")
def home_top_price_changes():
    """Home — 최신 고시일 vs 직전 고시일 약가 변동 Top N.
    abs(delta_pct) 기준 정렬. 변동사유는 reason_cache 에 있을 때만 포함.
    """
    try:
        limit = int(request.args.get("limit", "10"))
        with db._connect() as conn:
            dates = [r[0] for r in conn.execute(
                "SELECT DISTINCT apply_date FROM drug_prices "
                "ORDER BY apply_date DESC LIMIT 2"
            ).fetchall()]
            if len(dates) < 2:
                return jsonify({"latest": None, "prev": None, "items": []})
            latest, prev = dates[0], dates[1]

            # 최신 고시일 레코드 중 직전에도 존재하고 가격이 변한 것만
            rows = conn.execute(
                """
                SELECT c.insurance_code, c.product_name_kr, c.ingredient,
                       c.company, c.dosage_form, c.max_price, p.max_price
                FROM drug_prices c
                JOIN drug_prices p
                  ON p.insurance_code = c.insurance_code AND p.apply_date = ?
                WHERE c.apply_date = ? AND c.max_price IS NOT NULL
                  AND p.max_price IS NOT NULL AND p.max_price > 0
                  AND c.max_price != p.max_price
                """,
                (prev, latest),
            ).fetchall()

        # 1) 개별 variant 수집
        variants = []
        for r in rows:
            curr, prev_price = int(r[5]), int(r[6])
            delta = curr - prev_price
            pct = (delta / prev_price) * 100
            parsed = _parse_product(r[1] or "")
            variants.append({
                "insurance_code": r[0],
                "product_name": r[1] or "",
                "brand_name": parsed["brand"],
                "ingredient": r[2] or parsed["ingredient"],
                "company": r[3] or "",
                "dosage_form": r[4] or parsed["dosage_form"],
                "prev_price": prev_price,
                "curr_price": curr,
                "delta": delta,
                "delta_pct": round(pct, 2),
            })

        # 2) 제품명 base (= brand_name) 으로 그룹화 — 동일 브랜드는 함량/포장 variant 합치기
        groups: dict[str, dict] = {}
        for v in variants:
            key = v["brand_name"] or v["product_name"]
            g = groups.get(key)
            if g is None:
                groups[key] = {
                    **v,
                    "variant_count": 1,
                    "variants": [v],
                }
            else:
                g["variant_count"] += 1
                g["variants"].append(v)
                # 대표 variant: 변동률 절댓값 최대
                if abs(v["delta_pct"]) > abs(g["delta_pct"]):
                    for k in ("insurance_code", "product_name", "ingredient", "company",
                             "dosage_form", "prev_price", "curr_price", "delta", "delta_pct"):
                        g[k] = v[k]

        # 3) 비고 ("외 N정") 생성 + 변동률 절댓값 내림차순 정렬
        items = []
        for g in groups.values():
            cnt = g["variant_count"]
            # 제형 단위 (정 / 주 / 캡슐 / 시럽 등) — 대표 variant 의 dosage_form 또는 product_name 에서 유추
            unit = _extract_dose_unit(g["product_name"]) or "건"
            remark = f"외 {cnt - 1}{unit}" if cnt > 1 else ""
            items.append({
                "insurance_code": g["insurance_code"],
                "product_name": g["product_name"],
                "brand_name": g["brand_name"],
                "ingredient": g["ingredient"],
                "company": g["company"],
                "dosage_form": g["dosage_form"],
                "prev_price": g["prev_price"],
                "curr_price": g["curr_price"],
                "delta": g["delta"],
                "delta_pct": g["delta_pct"],
                "variant_count": cnt,
                "remark": remark,
            })
        items.sort(key=lambda x: abs(x["delta_pct"]), reverse=True)
        items = items[:limit]

        return jsonify({
            "latest_apply_date": latest,
            "prev_apply_date": prev,
            "count": len(items),
            "items": items,
        })
    except Exception as e:
        logger.error("home top price changes 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/msd/reimbursed-products")
def msd_reimbursed_products():
    """MSD 급여 등재 품목 상세 목록 — 최신 고시일 기준.
    Home 카드 클릭 시 팝업용.
    """
    try:
        with db._connect() as conn:
            latest_row = conn.execute(
                "SELECT MAX(apply_date) FROM drug_prices"
            ).fetchone()
            latest_apply_date = latest_row[0] if latest_row and latest_row[0] else None
            if not latest_apply_date:
                return jsonify({"latest_apply_date": None, "count": 0, "items": []})

            rows = conn.execute(
                """
                SELECT insurance_code, product_name_kr, ingredient,
                       dosage_form, dosage_strength, max_price, coverage_start
                FROM drug_prices
                WHERE company LIKE ? AND apply_date = ? AND max_price > 0
                ORDER BY product_name_kr, dosage_strength
                """,
                ("%엠에스디%", latest_apply_date),
            ).fetchall()

            # product_name_kr 기준 중복 제거 (함량·포장 병합) + parse 로 성분/제형 추출
            seen: dict = {}
            for r in rows:
                key = r[1] or ""
                if key not in seen:
                    parsed = _parse_product(key)
                    seen[key] = {
                        "insurance_code": r[0],
                        "product_name": r[1],
                        "brand_name": parsed["brand"],
                        "ingredient": r[2] or parsed["ingredient"],
                        "dosage_form": r[3] or parsed["dosage_form"],
                        "dosage_strength": r[4] or "",
                        "max_price": int(r[5]) if r[5] else 0,
                        "coverage_start": r[6] or "",
                    }

            items = sorted(seen.values(), key=lambda x: x["product_name"])
            return jsonify({
                "latest_apply_date": latest_apply_date,
                "count": len(items),
                "items": items,
            })
    except Exception as e:
        logger.error("MSD reimbursed products 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# Market Share (국내 IQVIA NSA-E)
# ──────────────────────────────────────────────────────────────────────────────

def _ms_latest_quarter(conn) -> str | None:
    row = conn.execute(
        "SELECT quarter FROM market_share_quarterly ORDER BY quarter DESC LIMIT 1"
    ).fetchone()
    return row[0] if row else None


def _ms_all_quarters(conn) -> list[str]:
    rows = conn.execute(
        "SELECT DISTINCT quarter FROM market_share_quarterly ORDER BY quarter ASC"
    ).fetchall()
    return [r[0] for r in rows]


@app.get("/api/market-share/search")
def market_share_search():
    """제품명/성분명 검색 — brand 단위 (product_name+atc4_code) 그룹, 최신 분기 값 포함."""
    q = (request.args.get("q") or "").strip()
    limit = min(int(request.args.get("limit", 30)), 100)
    if not q:
        return jsonify({"quarter": None, "items": []})
    try:
        with db._connect() as conn:
            latest = _ms_latest_quarter(conn)
            like = f"%{q.lower()}%"
            sql = """
                SELECT p.product_name, p.molecule_desc, p.mfr_name,
                       p.atc4_code, p.atc4_desc,
                       COUNT(DISTINCT p.product_id) AS pack_count,
                       COALESCE(SUM(q.values_lc), 0) AS values_lc,
                       COALESCE(SUM(q.dosage_units), 0) AS dosage_units
                FROM market_share_products p
                LEFT JOIN market_share_quarterly q
                       ON q.product_id = p.product_id AND q.quarter = ?
                WHERE LOWER(p.product_name) LIKE ?
                   OR LOWER(p.molecule_desc) LIKE ?
                GROUP BY p.product_name, p.molecule_desc, p.mfr_name, p.atc4_code, p.atc4_desc
                ORDER BY values_lc DESC
                LIMIT ?
            """
            rows = conn.execute(sql, (latest, like, like, limit)).fetchall()

            items = [
                {
                    "product_name": r[0],
                    "molecule_desc": r[1],
                    "mfr_name": r[2],
                    "atc4_code": r[3],
                    "atc4_desc": r[4],
                    "pack_count": int(r[5] or 0),
                    "values_lc": float(r[6] or 0.0),
                    "dosage_units": float(r[7] or 0.0),
                }
                for r in rows
            ]
        return jsonify({"quarter": latest, "items": items})
    except Exception as e:
        logger.error("market_share_search 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/market-share/atc4/<atc4_code>")
def market_share_atc4(atc4_code: str):
    """ATC4 시장 — 특정 분기의 브랜드별 값/점유율."""
    quarter = request.args.get("quarter")
    try:
        with db._connect() as conn:
            quarters = _ms_all_quarters(conn)
            if not quarters:
                return jsonify({"error": "no data"}), 404
            if quarter not in quarters:
                quarter = quarters[-1]

            meta = conn.execute(
                "SELECT atc4_desc FROM market_share_products WHERE atc4_code=? LIMIT 1",
                (atc4_code,),
            ).fetchone()
            if not meta:
                return jsonify({"error": f"atc4 {atc4_code} 없음"}), 404
            atc4_desc = meta[0]

            rows = conn.execute(
                """
                SELECT p.product_name, p.molecule_desc, p.mfr_name,
                       COUNT(DISTINCT p.product_id) AS pack_count,
                       COALESCE(SUM(q.values_lc), 0) AS values_lc,
                       COALESCE(SUM(q.dosage_units), 0) AS dosage_units
                FROM market_share_products p
                LEFT JOIN market_share_quarterly q
                       ON q.product_id = p.product_id AND q.quarter = ?
                WHERE p.atc4_code = ?
                GROUP BY p.product_name, p.molecule_desc, p.mfr_name
                ORDER BY values_lc DESC
                """,
                (quarter, atc4_code),
            ).fetchall()

            total_values = sum(float(r[4] or 0.0) for r in rows)
            total_units = sum(float(r[5] or 0.0) for r in rows)

            products = []
            for r in rows:
                v = float(r[4] or 0.0)
                u = float(r[5] or 0.0)
                products.append({
                    "product_name": r[0],
                    "molecule_desc": r[1],
                    "mfr_name": r[2],
                    "pack_count": int(r[3] or 0),
                    "values_lc": v,
                    "dosage_units": u,
                    "values_share_pct": (v / total_values * 100.0) if total_values > 0 else 0.0,
                    "units_share_pct": (u / total_units * 100.0) if total_units > 0 else 0.0,
                })

        return jsonify({
            "atc4_code": atc4_code,
            "atc4_desc": atc4_desc,
            "quarter": quarter,
            "quarters": quarters,
            "totals": {"values_lc": total_values, "dosage_units": total_units},
            "products": products,
        })
    except Exception as e:
        logger.error("market_share_atc4 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/market-share/export")
def market_share_export():
    """ATC4 시장 — Market Share / Unit Trend / Revenue Trend 3개 시트 xlsx.
    GET /api/market-share/export?atc4=L01G5&quarter=2025Q1&top=8
    """
    atc4_code = (request.args.get("atc4") or "").strip()
    quarter_arg = (request.args.get("quarter") or "").strip()
    top_n = min(int(request.args.get("top", 8)), 15)
    if not atc4_code:
        return jsonify({"error": "atc4 필수"}), 400

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl 미설치"}), 500

    try:
        with db._connect() as conn:
            quarters = _ms_all_quarters(conn)
            if not quarters:
                return jsonify({"error": "no data"}), 404
            quarter = quarter_arg if quarter_arg in quarters else quarters[-1]

            meta = conn.execute(
                "SELECT atc4_desc FROM market_share_products WHERE atc4_code=? LIMIT 1",
                (atc4_code,),
            ).fetchone()
            atc4_desc = meta[0] if meta else atc4_code

            # Sheet 1: Market Share (선택 분기)
            share_rows = conn.execute(
                """
                SELECT p.product_name, p.molecule_desc, p.mfr_name,
                       COALESCE(SUM(q.values_lc), 0)  AS values_lc,
                       COALESCE(SUM(q.dosage_units), 0) AS units
                FROM market_share_products p
                LEFT JOIN market_share_quarterly q
                       ON q.product_id = p.product_id AND q.quarter = ?
                WHERE p.atc4_code = ?
                GROUP BY p.product_name, p.molecule_desc, p.mfr_name
                ORDER BY values_lc DESC
                """,
                (quarter, atc4_code),
            ).fetchall()
            tot_v = sum(float(r[3] or 0.0) for r in share_rows) or 1.0
            tot_u = sum(float(r[4] or 0.0) for r in share_rows) or 1.0

            # Sheet 2 & 3: Unit / Revenue Trend (top-N)
            top_rows = conn.execute(
                """
                SELECT p.product_name
                FROM market_share_products p
                LEFT JOIN market_share_quarterly q
                       ON q.product_id = p.product_id AND q.quarter = ?
                WHERE p.atc4_code = ?
                GROUP BY p.product_name
                ORDER BY COALESCE(SUM(q.values_lc), 0) DESC
                LIMIT ?
                """,
                (quarter, atc4_code, top_n),
            ).fetchall()
            top_brands = [r[0] for r in top_rows]

            trend_rows = conn.execute(
                """
                SELECT q.quarter, p.product_name,
                       COALESCE(SUM(q.values_lc), 0)   AS v,
                       COALESCE(SUM(q.dosage_units), 0) AS u
                FROM market_share_products p
                JOIN market_share_quarterly q ON q.product_id = p.product_id
                WHERE p.atc4_code = ?
                GROUP BY q.quarter, p.product_name
                """,
                (atc4_code,),
            ).fetchall()

            q_tot = conn.execute(
                """
                SELECT q.quarter,
                       COALESCE(SUM(q.values_lc),0)   AS v,
                       COALESCE(SUM(q.dosage_units),0) AS u
                FROM market_share_products p
                JOIN market_share_quarterly q ON q.product_id = p.product_id
                WHERE p.atc4_code = ?
                GROUP BY q.quarter
                """,
                (atc4_code,),
            ).fetchall()
            qu = {r[0]: float(r[2] or 0.0) for r in q_tot}

            unit_share: dict[str, dict[str, float]] = {b: {} for b in top_brands}
            rev_series: dict[str, dict[str, float]] = {b: {} for b in top_brands}
            for qtr, name, v, u in trend_rows:
                if name not in unit_share:
                    continue
                unit_share[name][qtr] = (float(u or 0.0) / qu[qtr] * 100.0) if qu.get(qtr) else 0.0
                rev_series[name][qtr]  = float(v or 0.0) / 1_000_000.0

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1A56DB")

        def _style_header(ws):
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        # ── Sheet 1: Market Share
        ws1 = wb.active
        ws1.title = "Market Share"
        ws1.append(["Product", "Molecule", "Manufacturer",
                    "Values LC", "Dosage Units",
                    "Values Share (%)", "Units Share (%)"])
        for pn, mol, mfr, v, u in share_rows:
            v = float(v or 0.0); u = float(u or 0.0)
            ws1.append([pn, mol, mfr,
                        round(v, 2), round(u, 2),
                        round(v / tot_v * 100.0, 2),
                        round(u / tot_u * 100.0, 2)])
        _style_header(ws1)

        # ── Sheet 2: Unit Trend (점유율 %)
        ws2 = wb.create_sheet("Unit Trend")
        ws2.append(["Quarter", *top_brands])
        for qtr in quarters:
            row = [qtr]
            for b in top_brands:
                row.append(round(unit_share[b].get(qtr, 0.0), 2))
            ws2.append(row)
        _style_header(ws2)

        # ── Sheet 3: Revenue Trend (M KRW)
        ws3 = wb.create_sheet("Revenue Trend")
        ws3.append(["Quarter", *[f"{b} (M KRW)" for b in top_brands]])
        for qtr in quarters:
            row = [qtr]
            for b in top_brands:
                row.append(round(rev_series[b].get(qtr, 0.0), 2))
            ws3.append(row)
        _style_header(ws3)

        # 폭 자동
        for ws in (ws1, ws2, ws3):
            for col_idx, cell in enumerate(ws[1], start=1):
                values = [str(r[col_idx - 1]) for r in ws.iter_rows(values_only=True)]
                max_len = max([len(v or "") for v in values] + [10]) + 2
                ws.column_dimensions[cell.column_letter].width = min(max_len, 30)

        import io
        from datetime import datetime
        from urllib.parse import quote
        from flask import Response
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"MarketShare_{atc4_code}_{atc4_desc}_{quarter}_{stamp}.xlsx"
        safe = re.sub(r"[^\w\-.]", "_", fname)[:120]
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(safe)}"},
        )
    except Exception as e:
        logger.error("market_share_export 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/market-share/atc4/<atc4_code>/trend")
def market_share_atc4_trend(atc4_code: str):
    """ATC4 시장 — 브랜드별 분기 트렌드 (top-N)."""
    top_n = min(int(request.args.get("top", 6)), 15)
    try:
        with db._connect() as conn:
            quarters = _ms_all_quarters(conn)
            if not quarters:
                return jsonify({"error": "no data"}), 404
            latest = quarters[-1]

            meta = conn.execute(
                "SELECT atc4_desc FROM market_share_products WHERE atc4_code=? LIMIT 1",
                (atc4_code,),
            ).fetchone()
            if not meta:
                return jsonify({"error": f"atc4 {atc4_code} 없음"}), 404
            atc4_desc = meta[0]

            top_rows = conn.execute(
                """
                SELECT p.product_name,
                       COALESCE(SUM(q.values_lc), 0) AS values_lc
                FROM market_share_products p
                LEFT JOIN market_share_quarterly q
                       ON q.product_id = p.product_id AND q.quarter = ?
                WHERE p.atc4_code = ?
                GROUP BY p.product_name
                ORDER BY values_lc DESC
                LIMIT ?
                """,
                (latest, atc4_code, top_n),
            ).fetchall()
            top_names = [r[0] for r in top_rows]

            # 분기 × 브랜드 집계 (values & units)
            trend_rows = conn.execute(
                """
                SELECT q.quarter, p.product_name,
                       SUM(q.values_lc) AS v,
                       SUM(q.dosage_units) AS u
                FROM market_share_products p
                JOIN market_share_quarterly q ON q.product_id = p.product_id
                WHERE p.atc4_code = ?
                GROUP BY q.quarter, p.product_name
                """,
                (atc4_code,),
            ).fetchall()

            # 분기별 전체 합계
            quarter_totals = conn.execute(
                """
                SELECT q.quarter,
                       SUM(q.values_lc) AS v,
                       SUM(q.dosage_units) AS u
                FROM market_share_products p
                JOIN market_share_quarterly q ON q.product_id = p.product_id
                WHERE p.atc4_code = ?
                GROUP BY q.quarter
                """,
                (atc4_code,),
            ).fetchall()
            tot_v = {r[0]: float(r[1] or 0.0) for r in quarter_totals}
            tot_u = {r[0]: float(r[2] or 0.0) for r in quarter_totals}

            series: dict[str, dict] = {
                n: {"values": {}, "units": {}, "values_share": {}, "units_share": {}}
                for n in top_names
            }
            for r in trend_rows:
                qtr, name, v, u = r[0], r[1], float(r[2] or 0.0), float(r[3] or 0.0)
                if name not in series:
                    continue
                series[name]["values"][qtr] = v
                series[name]["units"][qtr] = u
                series[name]["values_share"][qtr] = (v / tot_v[qtr] * 100.0) if tot_v.get(qtr) else 0.0
                series[name]["units_share"][qtr] = (u / tot_u[qtr] * 100.0) if tot_u.get(qtr) else 0.0

        return jsonify({
            "atc4_code": atc4_code,
            "atc4_desc": atc4_desc,
            "quarters": quarters,
            "top_brands": top_names,
            "series": series,
        })
    except Exception as e:
        logger.error("market_share_atc4_trend 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/market-share/brand")
def market_share_brand():
    """브랜드 상세 — (product_name, atc4_code) 기준. 패키지별 내역 + 분기 시계열."""
    name = (request.args.get("name") or "").strip()
    atc4 = (request.args.get("atc4") or "").strip()
    if not name or not atc4:
        return jsonify({"error": "name, atc4 쿼리 필수"}), 400
    try:
        with db._connect() as conn:
            quarters = _ms_all_quarters(conn)
            latest = quarters[-1] if quarters else None

            pack_rows = conn.execute(
                """
                SELECT product_id, product_name, molecule_desc, mfr_name, corp, mnc13,
                       pack, pack_desc, strength, pack_launch_date,
                       atc4_code, atc4_desc
                FROM market_share_products
                WHERE product_name = ? AND atc4_code = ?
                """,
                (name, atc4),
            ).fetchall()
            if not pack_rows:
                return jsonify({"error": "brand not found"}), 404

            pack_ids = [r[0] for r in pack_rows]
            placeholders = ",".join(["?"] * len(pack_ids))
            qrows = conn.execute(
                f"""
                SELECT quarter, SUM(values_lc), SUM(dosage_units)
                FROM market_share_quarterly
                WHERE product_id IN ({placeholders})
                GROUP BY quarter
                ORDER BY quarter ASC
                """,
                pack_ids,
            ).fetchall()
            quarterly = [
                {"quarter": r[0], "values_lc": float(r[1] or 0.0), "dosage_units": float(r[2] or 0.0)}
                for r in qrows
            ]

            # ATC4 내 순위 / 점유율 (최신 분기)
            rank_sql = """
                SELECT p.product_name, COALESCE(SUM(q.values_lc), 0) AS v
                FROM market_share_products p
                LEFT JOIN market_share_quarterly q
                       ON q.product_id = p.product_id AND q.quarter = ?
                WHERE p.atc4_code = ?
                GROUP BY p.product_name
                ORDER BY v DESC
            """
            ranked = conn.execute(rank_sql, (latest, atc4)).fetchall()
            total = sum(float(r[1] or 0.0) for r in ranked)
            rank = next((i for i, r in enumerate(ranked) if r[0] == name), -1) + 1
            brand_v = next((float(r[1] or 0.0) for r in ranked if r[0] == name), 0.0)
            share_pct = (brand_v / total * 100.0) if total > 0 else 0.0

            packs = [
                {
                    "product_id": r[0],
                    "pack": r[6],
                    "pack_desc": r[7],
                    "strength": r[8],
                    "pack_launch_date": r[9],
                }
                for r in pack_rows
            ]
            first = pack_rows[0]

        return jsonify({
            "product_name": first[1],
            "molecule_desc": first[2],
            "mfr_name": first[3],
            "corp": first[4],
            "mnc13": first[5],
            "atc4_code": first[10],
            "atc4_desc": first[11],
            "quarter": latest,
            "quarters": quarters,
            "market_rank": rank if rank > 0 else None,
            "market_share_pct": share_pct,
            "market_total_values_lc": total,
            "packs": packs,
            "quarterly": quarterly,
        })
    except Exception as e:
        logger.error("market_share_brand 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# Market Share 관리자 업로드
# ──────────────────────────────────────────────────────────────────────────────

import os  # noqa: E402
import tempfile  # noqa: E402
import json as _json  # noqa: E402


@app.post("/api/admin/market-share/upload")
@require_auth(role="admin")
def admin_market_share_upload():
    """IQVIA NSA-E 분기 Excel 업로드 → 적재."""
    if "file" not in request.files:
        return jsonify({"error": "file 필드 누락", "code": "NO_FILE"}), 400
    file = request.files["file"]
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "xlsx 파일만 허용", "code": "BAD_EXT"}), 400
    sheet = request.form.get("sheet") or "NSA"
    user_email = getattr(request, "user", {}).get("sub") if hasattr(request, "user") else None

    tmp_path: Path | None = None
    try:
        fd, tmp = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        tmp_path = Path(tmp)
        file.save(str(tmp_path))

        result = ingest_market_share(
            xlsx_path=tmp_path,
            db_path=db.db_path,
            sheet_name=sheet,
            uploaded_by=user_email,
        )
        result["filename"] = file.filename
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e), "code": "BAD_SHEET"}), 400
    except Exception as e:
        logger.error("market_share_upload 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e), "code": "INGEST_FAIL"}), 500
    finally:
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


@app.get("/api/admin/market-share/uploads")
@require_auth(role="admin")
def admin_market_share_uploads():
    """최근 업로드 이력 (최대 50건)."""
    try:
        with db._connect() as conn:
            rows = conn.execute(
                """
                SELECT id, uploaded_at, uploaded_by, filename, rows_ingested, quarters_json
                FROM market_share_upload_log
                ORDER BY id DESC LIMIT 50
                """
            ).fetchall()
            total_products = conn.execute(
                "SELECT COUNT(*) FROM market_share_products"
            ).fetchone()[0]
            total_points = conn.execute(
                "SELECT COUNT(*) FROM market_share_quarterly"
            ).fetchone()[0]
            quarters = _ms_all_quarters(conn)
        uploads = [
            {
                "id": r[0],
                "uploaded_at": r[1],
                "uploaded_by": r[2],
                "filename": r[3],
                "rows_ingested": r[4],
                "quarters": _json.loads(r[5]) if r[5] else [],
            }
            for r in rows
        ]
        return jsonify({
            "uploads": uploads,
            "totals": {
                "products": total_products,
                "quarterly_points": total_points,
                "quarters_available": quarters,
            },
        })
    except Exception as e:
        logger.error("market_share_uploads 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# MSD Korea 파이프라인
# ──────────────────────────────────────────────────────────────────────────────

def _pipeline_row_to_dict(r) -> dict:
    return {
        "id": r[0],
        "name": r[1],
        "phase": r[2],
        "indication": r[3],
        "expected_year": r[4],
        "status": r[5],
        "note": r[6],
        "created_at": r[7],
        "updated_at": r[8],
    }


@app.get("/api/msd/pipeline")
@require_auth()
def msd_pipeline_list():
    """전체 파이프라인 (인증 사용자 공통). status/year 로 필터 가능."""
    status = request.args.get("status")
    year = request.args.get("year")
    try:
        with db._connect() as conn:
            sql = "SELECT id, name, phase, indication, expected_year, status, note, created_at, updated_at FROM msd_pipeline"
            params: list = []
            conds: list[str] = []
            if status in ("current", "upcoming"):
                conds.append("status = ?")
                params.append(status)
            if year:
                try:
                    conds.append("expected_year = ?")
                    params.append(int(year))
                except ValueError:
                    pass
            if conds:
                sql += " WHERE " + " AND ".join(conds)
            sql += " ORDER BY expected_year ASC, name ASC"
            rows = conn.execute(sql, params).fetchall()
        return jsonify({"items": [_pipeline_row_to_dict(r) for r in rows]})
    except Exception as e:
        logger.error("msd_pipeline_list 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.post("/api/admin/msd/pipeline")
@require_auth(role="admin")
def msd_pipeline_create():
    body = request.get_json(silent=True) or {}
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name 필수", "code": "INVALID"}), 400
    phase = body.get("phase") or None
    indication = body.get("indication") or None
    year = body.get("expected_year")
    try:
        year_int = int(year) if year not in (None, "") else None
    except (TypeError, ValueError):
        return jsonify({"error": "expected_year must be integer", "code": "INVALID"}), 400
    status = body.get("status") or "upcoming"
    if status not in ("current", "upcoming"):
        return jsonify({"error": "status must be current|upcoming", "code": "INVALID"}), 400
    note = body.get("note") or None
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc).isoformat()
    try:
        with db._connect() as conn:
            cur = conn.execute(
                "INSERT INTO msd_pipeline (name, phase, indication, expected_year, status, note, created_at, updated_at) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (name, phase, indication, year_int, status, note, now, now),
            )
            new_id = cur.lastrowid
            conn.commit()
            row = conn.execute(
                "SELECT id, name, phase, indication, expected_year, status, note, created_at, updated_at "
                "FROM msd_pipeline WHERE id=?",
                (new_id,),
            ).fetchone()
        return jsonify({"item": _pipeline_row_to_dict(row)}), 201
    except Exception as e:
        logger.error("msd_pipeline_create 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.patch("/api/admin/msd/pipeline/<int:item_id>")
@require_auth(role="admin")
def msd_pipeline_update(item_id: int):
    body = request.get_json(silent=True) or {}
    updatable = {"name", "phase", "indication", "expected_year", "status", "note"}
    fields = {k: v for k, v in body.items() if k in updatable}
    if not fields:
        return jsonify({"error": "변경할 필드 없음", "code": "INVALID"}), 400
    if "status" in fields and fields["status"] not in ("current", "upcoming"):
        return jsonify({"error": "status must be current|upcoming", "code": "INVALID"}), 400
    if "expected_year" in fields and fields["expected_year"] not in (None, ""):
        try:
            fields["expected_year"] = int(fields["expected_year"])
        except (TypeError, ValueError):
            return jsonify({"error": "expected_year must be integer", "code": "INVALID"}), 400
    elif "expected_year" in fields:
        fields["expected_year"] = None
    from datetime import datetime, timezone
    fields["updated_at"] = datetime.now(timezone.utc).isoformat()
    set_clause = ", ".join(f"{k} = ?" for k in fields.keys())
    params = list(fields.values()) + [item_id]
    try:
        with db._connect() as conn:
            res = conn.execute(
                f"UPDATE msd_pipeline SET {set_clause} WHERE id = ?", params
            )
            if res.rowcount == 0:
                return jsonify({"error": "not found", "code": "NOT_FOUND"}), 404
            conn.commit()
            row = conn.execute(
                "SELECT id, name, phase, indication, expected_year, status, note, created_at, updated_at "
                "FROM msd_pipeline WHERE id=?",
                (item_id,),
            ).fetchone()
        return jsonify({"item": _pipeline_row_to_dict(row)})
    except Exception as e:
        logger.error("msd_pipeline_update 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.delete("/api/admin/msd/pipeline/<int:item_id>")
@require_auth(role="admin")
def msd_pipeline_delete(item_id: int):
    try:
        with db._connect() as conn:
            res = conn.execute("DELETE FROM msd_pipeline WHERE id = ?", (item_id,))
            if res.rowcount == 0:
                return jsonify({"error": "not found", "code": "NOT_FOUND"}), 404
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        logger.error("msd_pipeline_delete 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# 브랜드 미디어 트래픽 (Home — 브랜드 언급 Top 10)
# ──────────────────────────────────────────────────────────────────────────────

def _brand_traffic_row_to_dict(r) -> dict:
    import json as _json
    return {
        "id": r[0],
        "rank": r[1],
        "brand": r[2],
        "company": r[3],
        "category": r[4],
        "color": r[5],
        "trafficIndex": r[6],
        "change": r[7],
        "sparkline": _json.loads(r[8]) if r[8] else [],
        "news": _json.loads(r[9]) if r[9] else [],
        "created_at": r[10],
        "updated_at": r[11],
    }


_BRAND_TRAFFIC_COLS = "id, rank, brand, company, category, color, traffic_index, change_pct, sparkline_json, news_json, created_at, updated_at"


@app.get("/api/brand-traffic")
@require_auth()
def brand_traffic_list():
    try:
        with db._connect() as conn:
            rows = conn.execute(
                f"SELECT {_BRAND_TRAFFIC_COLS} FROM brand_traffic ORDER BY rank ASC"
            ).fetchall()
        return jsonify({"items": [_brand_traffic_row_to_dict(r) for r in rows]})
    except Exception as e:
        logger.error("brand_traffic_list 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


def _coerce_brand_traffic_input(body: dict) -> dict | tuple[dict, str]:
    import json as _json
    out: dict = {}
    if "rank" in body:
        try:
            out["rank"] = int(body["rank"])
        except (TypeError, ValueError):
            return ({}, "rank must be integer")
    if "brand" in body:
        brand = (body.get("brand") or "").strip()
        if not brand:
            return ({}, "brand required")
        out["brand"] = brand
    for k in ("company", "category", "color"):
        if k in body:
            v = body.get(k)
            out[k] = v.strip() if isinstance(v, str) else v
    if "trafficIndex" in body:
        try:
            out["traffic_index"] = int(body["trafficIndex"])
        except (TypeError, ValueError):
            return ({}, "trafficIndex must be integer")
    if "change" in body:
        try:
            out["change_pct"] = float(body["change"])
        except (TypeError, ValueError):
            return ({}, "change must be number")
    if "sparkline" in body:
        spark = body.get("sparkline") or []
        if not isinstance(spark, list):
            return ({}, "sparkline must be array")
        try:
            spark = [float(x) for x in spark]
        except (TypeError, ValueError):
            return ({}, "sparkline must contain numbers")
        out["sparkline_json"] = _json.dumps(spark, ensure_ascii=False)
    if "news" in body:
        news = body.get("news") or []
        if not isinstance(news, list):
            return ({}, "news must be array")
        out["news_json"] = _json.dumps(news, ensure_ascii=False)
    return out


@app.post("/api/admin/brand-traffic")
@require_auth(role="admin")
def brand_traffic_create():
    body = request.get_json(silent=True) or {}
    if not (body.get("brand") or "").strip():
        return jsonify({"error": "brand required", "code": "INVALID"}), 400
    result = _coerce_brand_traffic_input(body)
    if isinstance(result, tuple):
        _, msg = result
        return jsonify({"error": msg, "code": "INVALID"}), 400
    fields = result
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc).isoformat()
    try:
        with db._connect() as conn:
            if "rank" not in fields:
                row = conn.execute("SELECT COALESCE(MAX(rank),0)+1 FROM brand_traffic").fetchone()
                fields["rank"] = row[0]
            fields.setdefault("traffic_index", 0)
            fields.setdefault("change_pct", 0.0)
            fields.setdefault("sparkline_json", "[]")
            fields.setdefault("news_json", "[]")
            fields["created_at"] = now
            fields["updated_at"] = now
            cols = ", ".join(fields.keys())
            placeholders = ", ".join("?" for _ in fields)
            cur = conn.execute(
                f"INSERT INTO brand_traffic ({cols}) VALUES ({placeholders})",
                list(fields.values()),
            )
            conn.commit()
            new_id = cur.lastrowid
            row = conn.execute(
                f"SELECT {_BRAND_TRAFFIC_COLS} FROM brand_traffic WHERE id=?", (new_id,)
            ).fetchone()
        return jsonify({"item": _brand_traffic_row_to_dict(row)}), 201
    except Exception as e:
        logger.error("brand_traffic_create 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.patch("/api/admin/brand-traffic/<int:item_id>")
@require_auth(role="admin")
def brand_traffic_update(item_id: int):
    body = request.get_json(silent=True) or {}
    result = _coerce_brand_traffic_input(body)
    if isinstance(result, tuple):
        _, msg = result
        return jsonify({"error": msg, "code": "INVALID"}), 400
    fields = result
    if not fields:
        return jsonify({"error": "변경할 필드 없음", "code": "INVALID"}), 400
    from datetime import datetime, timezone
    fields["updated_at"] = datetime.now(timezone.utc).isoformat()
    set_clause = ", ".join(f"{k} = ?" for k in fields.keys())
    params = list(fields.values()) + [item_id]
    try:
        with db._connect() as conn:
            res = conn.execute(
                f"UPDATE brand_traffic SET {set_clause} WHERE id = ?", params
            )
            if res.rowcount == 0:
                return jsonify({"error": "not found", "code": "NOT_FOUND"}), 404
            conn.commit()
            row = conn.execute(
                f"SELECT {_BRAND_TRAFFIC_COLS} FROM brand_traffic WHERE id=?", (item_id,)
            ).fetchone()
        return jsonify({"item": _brand_traffic_row_to_dict(row)})
    except Exception as e:
        logger.error("brand_traffic_update 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.delete("/api/admin/brand-traffic/<int:item_id>")
@require_auth(role="admin")
def brand_traffic_delete(item_id: int):
    try:
        with db._connect() as conn:
            res = conn.execute("DELETE FROM brand_traffic WHERE id = ?", (item_id,))
            if res.rowcount == 0:
                return jsonify({"error": "not found", "code": "NOT_FOUND"}), 404
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        logger.error("brand_traffic_delete 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# Daily Mailing — 구독 설정 (사용자별 CRUD + 테스트 발송)
# ──────────────────────────────────────────────────────────────────────────────

from agents.notify.mailer import (  # noqa: E402
    send_email as _mail_send,
    smtp_configured as _mail_smtp_configured,
)
from agents.notify.digest import render_daily_digest as _mail_render_digest  # noqa: E402


def _mail_sub_row_to_dict(r) -> dict:
    import json as _json
    return {
        "id": r[0],
        "name": r[1],
        "keywords": _json.loads(r[2]) if r[2] else [],
        "media": _json.loads(r[3]) if r[3] else [],
        "schedule": r[4],
        "time": r[5],
        "weekDay": r[6],
        "emails": _json.loads(r[7]) if r[7] else [],
        "active": bool(r[8]),
        "created_at": r[9],
        "updated_at": r[10],
        "last_sent_at": r[11],
    }


_MAIL_SUB_COLS = "id, name, keywords_json, media_json, schedule, time, week_day, emails_json, active, created_at, updated_at, last_sent_at"


def _coerce_mail_sub_input(body: dict) -> dict | tuple[dict, str]:
    import json as _json
    out: dict = {}
    if "name" in body:
        name = (body.get("name") or "").strip()
        if not name:
            return ({}, "name required")
        out["name"] = name
    if "keywords" in body:
        kws = body.get("keywords") or []
        if not isinstance(kws, list):
            return ({}, "keywords must be array")
        out["keywords_json"] = _json.dumps([str(x) for x in kws], ensure_ascii=False)
    if "media" in body:
        media = body.get("media") or []
        if not isinstance(media, list):
            return ({}, "media must be array")
        out["media_json"] = _json.dumps([str(x) for x in media], ensure_ascii=False)
    if "schedule" in body:
        sched = body.get("schedule")
        if sched not in ("Daily", "Weekly"):
            return ({}, "schedule must be Daily|Weekly")
        out["schedule"] = sched
    if "time" in body:
        t = (body.get("time") or "").strip()
        if not t:
            return ({}, "time required")
        out["time"] = t
    if "weekDay" in body:
        wd = body.get("weekDay")
        out["week_day"] = wd if wd else None
    if "emails" in body:
        emails = body.get("emails") or []
        if not isinstance(emails, list):
            return ({}, "emails must be array")
        cleaned = [e.strip() for e in emails if isinstance(e, str) and e.strip()]
        if not cleaned:
            return ({}, "at least one email required")
        out["emails_json"] = _json.dumps(cleaned, ensure_ascii=False)
    if "active" in body:
        out["active"] = 1 if body.get("active") else 0
    return out


@app.get("/api/mail-subscriptions")
@require_auth()
def mail_sub_list():
    owner = request.user["sub"]  # type: ignore[attr-defined]
    try:
        with db._connect() as conn:
            rows = conn.execute(
                f"SELECT {_MAIL_SUB_COLS} FROM mail_subscription WHERE owner_email=? ORDER BY created_at DESC",
                (owner,),
            ).fetchall()
        return jsonify({
            "items": [_mail_sub_row_to_dict(r) for r in rows],
            "smtp_configured": _mail_smtp_configured(),
        })
    except Exception as e:
        logger.error("mail_sub_list 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.post("/api/mail-subscriptions")
@require_auth()
def mail_sub_create():
    owner = request.user["sub"]  # type: ignore[attr-defined]
    body = request.get_json(silent=True) or {}
    required = ("name", "keywords", "media", "schedule", "time", "emails")
    for k in required:
        if k not in body:
            return jsonify({"error": f"{k} required", "code": "INVALID"}), 400
    result = _coerce_mail_sub_input(body)
    if isinstance(result, tuple):
        _, msg = result
        return jsonify({"error": msg, "code": "INVALID"}), 400
    fields = result
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc).isoformat()
    fields["owner_email"] = owner
    fields.setdefault("active", 1)
    fields.setdefault("week_day", None)
    fields["created_at"] = now
    fields["updated_at"] = now
    try:
        with db._connect() as conn:
            cols = ", ".join(fields.keys())
            placeholders = ", ".join("?" for _ in fields)
            cur = conn.execute(
                f"INSERT INTO mail_subscription ({cols}) VALUES ({placeholders})",
                list(fields.values()),
            )
            conn.commit()
            row = conn.execute(
                f"SELECT {_MAIL_SUB_COLS} FROM mail_subscription WHERE id=?", (cur.lastrowid,),
            ).fetchone()
        return jsonify({"item": _mail_sub_row_to_dict(row)}), 201
    except Exception as e:
        logger.error("mail_sub_create 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.patch("/api/mail-subscriptions/<int:item_id>")
@require_auth()
def mail_sub_update(item_id: int):
    owner = request.user["sub"]  # type: ignore[attr-defined]
    body = request.get_json(silent=True) or {}
    result = _coerce_mail_sub_input(body)
    if isinstance(result, tuple):
        _, msg = result
        return jsonify({"error": msg, "code": "INVALID"}), 400
    fields = result
    if not fields:
        return jsonify({"error": "변경할 필드 없음", "code": "INVALID"}), 400
    from datetime import datetime, timezone
    fields["updated_at"] = datetime.now(timezone.utc).isoformat()
    set_clause = ", ".join(f"{k} = ?" for k in fields.keys())
    params = list(fields.values()) + [item_id, owner]
    try:
        with db._connect() as conn:
            res = conn.execute(
                f"UPDATE mail_subscription SET {set_clause} WHERE id = ? AND owner_email = ?",
                params,
            )
            if res.rowcount == 0:
                return jsonify({"error": "not found", "code": "NOT_FOUND"}), 404
            conn.commit()
            row = conn.execute(
                f"SELECT {_MAIL_SUB_COLS} FROM mail_subscription WHERE id=?", (item_id,),
            ).fetchone()
        return jsonify({"item": _mail_sub_row_to_dict(row)})
    except Exception as e:
        logger.error("mail_sub_update 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.delete("/api/mail-subscriptions/<int:item_id>")
@require_auth()
def mail_sub_delete(item_id: int):
    owner = request.user["sub"]  # type: ignore[attr-defined]
    try:
        with db._connect() as conn:
            res = conn.execute(
                "DELETE FROM mail_subscription WHERE id = ? AND owner_email = ?",
                (item_id, owner),
            )
            if res.rowcount == 0:
                return jsonify({"error": "not found", "code": "NOT_FOUND"}), 404
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        logger.error("mail_sub_delete 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/mailing/preview")
def mail_preview():
    """Daily Mailing 프리뷰 — 실데이터로 렌더된 HTML 반환.

    Query:
      - name: 구독 이름 (default "Daily Dossier")
      - keywords: 콤마 구분
      - media: 콤마 구분
      - format: "html" (default) → text/html, "json" → {subject, html, text}
    """
    name = (request.args.get("name") or "Daily Dossier").strip()
    keywords = [s for s in (request.args.get("keywords") or "").split(",") if s.strip()]
    media = [s for s in (request.args.get("media") or "").split(",") if s.strip()]
    fmt = request.args.get("format", "html")
    dashboard_url = request.host_url.rstrip("/").replace(":5001", ":3000")
    try:
        subject, body_html, body_text = _mail_render_digest(
            name=name, dashboard_url=dashboard_url, keywords=keywords, media=media,
        )
        if fmt == "json":
            return jsonify({"subject": subject, "html": body_html, "text": body_text})
        from flask import Response
        return Response(body_html, mimetype="text/html")
    except Exception as e:
        logger.error("mail preview 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.post("/api/mail-subscriptions/<int:item_id>/preview")
@require_auth()
def mail_sub_preview(item_id: int):
    """저장된 구독에 대한 프리뷰 JSON ({subject, html, text})."""
    owner = request.user["sub"]  # type: ignore[attr-defined]
    try:
        with db._connect() as conn:
            row = conn.execute(
                f"SELECT {_MAIL_SUB_COLS} FROM mail_subscription WHERE id=? AND owner_email=?",
                (item_id, owner),
            ).fetchone()
        if row is None:
            return jsonify({"error": "not found", "code": "NOT_FOUND"}), 404
        item = _mail_sub_row_to_dict(row)
        dashboard_url = request.host_url.rstrip("/").replace(":5001", ":3000")
        subject, body_html, body_text = _mail_render_digest(
            name=item["name"], dashboard_url=dashboard_url,
            keywords=item["keywords"], media=item["media"],
        )
        return jsonify({"subject": subject, "html": body_html, "text": body_text})
    except Exception as e:
        logger.error("mail sub preview 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.post("/api/mail-subscriptions/<int:item_id>/test-send")
@require_auth()
def mail_sub_test_send(item_id: int):
    owner = request.user["sub"]  # type: ignore[attr-defined]
    try:
        with db._connect() as conn:
            row = conn.execute(
                f"SELECT {_MAIL_SUB_COLS} FROM mail_subscription WHERE id=? AND owner_email=?",
                (item_id, owner),
            ).fetchone()
            if row is None:
                return jsonify({"error": "not found", "code": "NOT_FOUND"}), 404
            item = _mail_sub_row_to_dict(row)
            dashboard_url = request.host_url.rstrip("/").replace(":5001", ":3000")
            subject, body_html, body_text = _mail_render_digest(
                name=item["name"],
                dashboard_url=dashboard_url,
                keywords=item["keywords"],
                media=item["media"],
            )
            result = _mail_send(item["emails"], subject, body_html, body_text=body_text)
            if result.get("ok") and result.get("mode") == "smtp":
                from datetime import datetime, timezone
                now = datetime.now(timezone.utc).isoformat()
                conn.execute(
                    "UPDATE mail_subscription SET last_sent_at=? WHERE id=?",
                    (now, item_id),
                )
                conn.commit()
        return jsonify(result)
    except Exception as e:
        logger.error("mail_sub_test_send 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# Competitor Trends — 경쟁사 동향 카드 (CRUD)
# ──────────────────────────────────────────────────────────────────────────────

_COMPETITOR_COLS = "id, company, logo, color, badge, badge_color, headline, detail, date, source, url, created_at, updated_at"

_COMPETITOR_BADGES = ("신규 출시", "가격 변동", "임상 진행", "급여 등재", "파이프라인", "전략 변화")


def _competitor_row_to_dict(r) -> dict:
    return {
        "id": r[0],
        "company": r[1],
        "logo": r[2],
        "color": r[3],
        "badge": r[4],
        "badgeColor": r[5],
        "headline": r[6],
        "detail": r[7],
        "date": r[8],
        "source": r[9],
        "url": r[10],
        "created_at": r[11],
        "updated_at": r[12],
    }


def _coerce_competitor_input(body: dict) -> dict | tuple[dict, str]:
    out: dict = {}
    for k in ("company", "headline", "detail"):
        if k in body:
            v = (body.get(k) or "").strip()
            if not v:
                return ({}, f"{k} required")
            out[k] = v
    if "badge" in body:
        badge = (body.get("badge") or "").strip()
        if badge not in _COMPETITOR_BADGES:
            return ({}, f"badge must be one of {_COMPETITOR_BADGES}")
        out["badge"] = badge
    if "date" in body:
        d = (body.get("date") or "").strip()
        if not d:
            return ({}, "date required")
        out["date"] = d
    for k in ("logo", "color", "source", "url"):
        if k in body:
            v = body.get(k)
            out[k] = v.strip() if isinstance(v, str) else v
    if "badgeColor" in body:
        v = body.get("badgeColor")
        out["badge_color"] = v.strip() if isinstance(v, str) else v
    return out


@app.get("/api/competitor-trends")
@require_auth()
def competitor_list():
    try:
        with db._connect() as conn:
            rows = conn.execute(
                f"SELECT {_COMPETITOR_COLS} FROM competitor_trend ORDER BY date DESC, id DESC"
            ).fetchall()
        return jsonify({"items": [_competitor_row_to_dict(r) for r in rows]})
    except Exception as e:
        logger.error("competitor_list 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.post("/api/admin/competitor-trends")
@require_auth(role="admin")
def competitor_create():
    body = request.get_json(silent=True) or {}
    for req_key in ("company", "badge", "headline", "detail", "date"):
        if not (body.get(req_key) or "").strip() if isinstance(body.get(req_key), str) else not body.get(req_key):
            return jsonify({"error": f"{req_key} required", "code": "INVALID"}), 400
    result = _coerce_competitor_input(body)
    if isinstance(result, tuple):
        _, msg = result
        return jsonify({"error": msg, "code": "INVALID"}), 400
    fields = result
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc).isoformat()
    fields["created_at"] = now
    fields["updated_at"] = now
    cols = ", ".join(fields.keys())
    placeholders = ", ".join("?" for _ in fields)
    try:
        with db._connect() as conn:
            cur = conn.execute(
                f"INSERT INTO competitor_trend ({cols}) VALUES ({placeholders})",
                list(fields.values()),
            )
            conn.commit()
            new_id = cur.lastrowid
            row = conn.execute(
                f"SELECT {_COMPETITOR_COLS} FROM competitor_trend WHERE id=?", (new_id,)
            ).fetchone()
        return jsonify({"item": _competitor_row_to_dict(row)}), 201
    except Exception as e:
        logger.error("competitor_create 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.patch("/api/admin/competitor-trends/<int:item_id>")
@require_auth(role="admin")
def competitor_update(item_id: int):
    body = request.get_json(silent=True) or {}
    result = _coerce_competitor_input(body)
    if isinstance(result, tuple):
        _, msg = result
        return jsonify({"error": msg, "code": "INVALID"}), 400
    fields = result
    if not fields:
        return jsonify({"error": "변경할 필드 없음", "code": "INVALID"}), 400
    from datetime import datetime, timezone
    fields["updated_at"] = datetime.now(timezone.utc).isoformat()
    set_clause = ", ".join(f"{k} = ?" for k in fields.keys())
    params = list(fields.values()) + [item_id]
    try:
        with db._connect() as conn:
            res = conn.execute(
                f"UPDATE competitor_trend SET {set_clause} WHERE id = ?", params
            )
            if res.rowcount == 0:
                return jsonify({"error": "not found", "code": "NOT_FOUND"}), 404
            conn.commit()
            row = conn.execute(
                f"SELECT {_COMPETITOR_COLS} FROM competitor_trend WHERE id=?", (item_id,)
            ).fetchone()
        return jsonify({"item": _competitor_row_to_dict(row)})
    except Exception as e:
        logger.error("competitor_update 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.delete("/api/admin/competitor-trends/<int:item_id>")
@require_auth(role="admin")
def competitor_delete(item_id: int):
    try:
        with db._connect() as conn:
            res = conn.execute("DELETE FROM competitor_trend WHERE id = ?", (item_id,))
            if res.rowcount == 0:
                return jsonify({"error": "not found", "code": "NOT_FOUND"}), 404
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        logger.error("competitor_delete 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.post("/api/admin/competitor-trends/refresh")
@require_auth(role="admin")
def competitor_refresh():
    """경쟁사 뉴스 수동 크롤 트리거 (주 1회 cron 과 동일 로직)."""
    body = request.get_json(silent=True) or {}
    days = int(body.get("days", 7))
    dry_run = bool(body.get("dry_run", False))
    model = body.get("model") or "gpt-4o-mini"
    try:
        from agents.competitor_trends_agent import run as _ct_run
        result = _ct_run(days=days, dry_run=dry_run, model=model)
        return jsonify(result)
    except Exception as e:
        logger.error("competitor_refresh 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e), "code": "REFRESH_FAIL"}), 500


# ──────────────────────────────────────────────────────────────────────────────
# Keyword Cloud — Home 워드클라우드 (CRUD)
# ──────────────────────────────────────────────────────────────────────────────

_KEYWORD_COLS = "id, text, weight, color, created_at, updated_at"


def _keyword_row_to_dict(r) -> dict:
    return {
        "id": r[0],
        "text": r[1],
        "weight": r[2],
        "color": r[3],
        "created_at": r[4],
        "updated_at": r[5],
    }


def _coerce_keyword_input(body: dict) -> dict | tuple[dict, str]:
    out: dict = {}
    if "text" in body:
        t = (body.get("text") or "").strip()
        if not t:
            return ({}, "text required")
        out["text"] = t
    if "weight" in body:
        try:
            w = int(body["weight"])
        except (TypeError, ValueError):
            return ({}, "weight must be integer")
        if w < 0 or w > 1000:
            return ({}, "weight out of range")
        out["weight"] = w
    if "color" in body:
        v = body.get("color")
        out["color"] = v.strip() if isinstance(v, str) else v
    return out


@app.get("/api/keyword-cloud")
@require_auth()
def keyword_list():
    try:
        with db._connect() as conn:
            rows = conn.execute(
                f"SELECT {_KEYWORD_COLS} FROM keyword_cloud ORDER BY weight DESC, id ASC"
            ).fetchall()
        return jsonify({"items": [_keyword_row_to_dict(r) for r in rows]})
    except Exception as e:
        logger.error("keyword_list 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.post("/api/admin/keyword-cloud")
@require_auth(role="admin")
def keyword_create():
    body = request.get_json(silent=True) or {}
    if not (body.get("text") or "").strip():
        return jsonify({"error": "text required", "code": "INVALID"}), 400
    result = _coerce_keyword_input(body)
    if isinstance(result, tuple):
        _, msg = result
        return jsonify({"error": msg, "code": "INVALID"}), 400
    fields = result
    fields.setdefault("weight", 50)
    fields.setdefault("color", "#8B9BB4")
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc).isoformat()
    fields["created_at"] = now
    fields["updated_at"] = now
    cols = ", ".join(fields.keys())
    placeholders = ", ".join("?" for _ in fields)
    try:
        with db._connect() as conn:
            try:
                cur = conn.execute(
                    f"INSERT INTO keyword_cloud ({cols}) VALUES ({placeholders})",
                    list(fields.values()),
                )
            except sqlite3.IntegrityError:
                return jsonify({"error": "duplicate text", "code": "CONFLICT"}), 409
            conn.commit()
            new_id = cur.lastrowid
            row = conn.execute(
                f"SELECT {_KEYWORD_COLS} FROM keyword_cloud WHERE id=?", (new_id,)
            ).fetchone()
        return jsonify({"item": _keyword_row_to_dict(row)}), 201
    except Exception as e:
        logger.error("keyword_create 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.patch("/api/admin/keyword-cloud/<int:item_id>")
@require_auth(role="admin")
def keyword_update(item_id: int):
    body = request.get_json(silent=True) or {}
    result = _coerce_keyword_input(body)
    if isinstance(result, tuple):
        _, msg = result
        return jsonify({"error": msg, "code": "INVALID"}), 400
    fields = result
    if not fields:
        return jsonify({"error": "변경할 필드 없음", "code": "INVALID"}), 400
    from datetime import datetime, timezone
    fields["updated_at"] = datetime.now(timezone.utc).isoformat()
    set_clause = ", ".join(f"{k} = ?" for k in fields.keys())
    params = list(fields.values()) + [item_id]
    try:
        with db._connect() as conn:
            res = conn.execute(
                f"UPDATE keyword_cloud SET {set_clause} WHERE id = ?", params
            )
            if res.rowcount == 0:
                return jsonify({"error": "not found", "code": "NOT_FOUND"}), 404
            conn.commit()
            row = conn.execute(
                f"SELECT {_KEYWORD_COLS} FROM keyword_cloud WHERE id=?", (item_id,)
            ).fetchone()
        return jsonify({"item": _keyword_row_to_dict(row)})
    except Exception as e:
        logger.error("keyword_update 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.delete("/api/admin/keyword-cloud/<int:item_id>")
@require_auth(role="admin")
def keyword_delete(item_id: int):
    try:
        with db._connect() as conn:
            res = conn.execute("DELETE FROM keyword_cloud WHERE id = ?", (item_id,))
            if res.rowcount == 0:
                return jsonify({"error": "not found", "code": "NOT_FOUND"}), 404
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        logger.error("keyword_delete 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# 적응증별 한국 급여 상태 (HIRA) — admin 체크리스트 CRUD
# ──────────────────────────────────────────────────────────────────────────────

def _reimbursement_row_to_dict(r) -> dict:
    return {
        "indication_id": r[0],
        "product": r[1],
        "disease": r[2],
        "line_of_therapy": r[3],
        "stage": r[4],
        "biomarker_class": r[5],
        "title": r[6],
        "is_reimbursed": bool(r[7]) if r[7] is not None else False,
        "effective_date": r[8],
        "criteria_text": r[9],
        "notice_date": r[10],
        "notice_url": r[11],
        "updated_by": r[12],
        "updated_at": r[13],
    }


@app.get("/api/admin/reimbursement")
@require_auth(role="admin")
def admin_reimbursement_list():
    """product 별 적응증 목록 + 급여 상태 (LEFT JOIN)."""
    product = (request.args.get("product") or "").strip().lower()
    try:
        with db._connect() as conn:
            sql = """
                SELECT m.indication_id, m.product, m.disease, m.line_of_therapy,
                       m.stage, m.biomarker_class, m.title,
                       r.is_reimbursed, r.effective_date, r.criteria_text,
                       r.notice_date, r.notice_url, r.updated_by, r.updated_at
                FROM indications_master m
                LEFT JOIN indication_reimbursement r
                    ON r.indication_id = m.indication_id
            """
            params: list = []
            if product:
                sql += " WHERE m.product = ?"
                params.append(product)
            sql += " ORDER BY m.product, m.disease, m.line_of_therapy, m.indication_id"
            rows = conn.execute(sql, params).fetchall()
        return jsonify({"items": [_reimbursement_row_to_dict(r) for r in rows]})
    except Exception as e:
        logger.error("reimbursement_list 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.put("/api/admin/reimbursement/<indication_id>")
@require_auth(role="admin")
def admin_reimbursement_upsert(indication_id: str):
    """indication_id 단위 upsert. body: {is_reimbursed, effective_date, criteria_text, notice_date, notice_url}."""
    body = request.get_json(silent=True) or {}
    is_reimbursed = 1 if body.get("is_reimbursed") else 0
    effective_date = (body.get("effective_date") or "").strip() or None
    criteria_text = (body.get("criteria_text") or "").strip() or None
    notice_date = (body.get("notice_date") or "").strip() or None
    notice_url = (body.get("notice_url") or "").strip() or None
    user_email = getattr(request, "user", {}).get("sub") if hasattr(request, "user") else None
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    try:
        with db._connect() as conn:
            exists = conn.execute(
                "SELECT 1 FROM indications_master WHERE indication_id = ?",
                (indication_id,),
            ).fetchone()
            if not exists:
                return jsonify({"error": "indication_id not found", "code": "NOT_FOUND"}), 404
            conn.execute(
                """
                INSERT INTO indication_reimbursement
                    (indication_id, is_reimbursed, effective_date, criteria_text,
                     notice_date, notice_url, updated_by, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(indication_id) DO UPDATE SET
                    is_reimbursed = excluded.is_reimbursed,
                    effective_date = excluded.effective_date,
                    criteria_text = excluded.criteria_text,
                    notice_date = excluded.notice_date,
                    notice_url = excluded.notice_url,
                    updated_by = excluded.updated_by,
                    updated_at = excluded.updated_at
                """,
                (indication_id, is_reimbursed, effective_date, criteria_text,
                 notice_date, notice_url, user_email, now),
            )
            conn.commit()
            row = conn.execute(
                """
                SELECT m.indication_id, m.product, m.disease, m.line_of_therapy,
                       m.stage, m.biomarker_class, m.title,
                       r.is_reimbursed, r.effective_date, r.criteria_text,
                       r.notice_date, r.notice_url, r.updated_by, r.updated_at
                FROM indications_master m
                LEFT JOIN indication_reimbursement r
                    ON r.indication_id = m.indication_id
                WHERE m.indication_id = ?
                """,
                (indication_id,),
            ).fetchone()
        return jsonify({"item": _reimbursement_row_to_dict(row)})
    except Exception as e:
        logger.error("reimbursement_upsert 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# 허가문서 (PDF) 업로드 — 5국 (EMA/MHRA/PMDA/TGA/MFDS) 수동 보강
# (FDA 자동 sync 는 /api/admin/fda-sync, indication grid 는 /api/admin/indication-grid)
# ──────────────────────────────────────────────────────────────────────────────

UPLOAD_DIR = BASE_DIR / "data" / "uploads" / "approval_pdf"


@app.get("/api/admin/indication-grid")
@require_auth(role="admin")
def admin_indication_grid():
    """product 별 적응증 × 6국 agency grid. UI 매트릭스 데이터.

    GET /api/admin/indication-grid?product=keytruda
    응답: {product, indications: [{indication_id, title, agencies: {FDA: {...}, EMA: {...}, ...}}]}
    """
    product = (request.args.get("product") or "").strip().lower()
    if not product:
        return jsonify({"error": "product 파라미터 필요"}), 400
    try:
        indications = db.get_approval_grid(product)
        return jsonify({"product": product, "indications": indications})
    except Exception as e:
        logger.error("indication_grid 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.post("/api/admin/approval-document")
@require_auth(role="admin")
def admin_approval_document_upload():
    """PDF 업로드 + indications_by_agency 병행 갱신.

    multipart/form-data:
      file: PDF
      indication_id: indications_master 의 PK
      agency: FDA / EMA / MHRA / PMDA / TGA / MFDS
      approval_date: YYYY-MM-DD (선택)
      label_excerpt: 적응증 본문 발췌 (선택)
      label_url: 원본 사이트 URL (선택)
      notes: 자유 메모 (선택)
    """
    if "file" not in request.files:
        return jsonify({"error": "file 필드 누락", "code": "NO_FILE"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "filename 누락", "code": "NO_NAME"}), 400
    if not file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "PDF 만 허용", "code": "BAD_EXT"}), 400

    indication_id = (request.form.get("indication_id") or "").strip()
    agency = (request.form.get("agency") or "").strip().upper()
    if not indication_id or not agency:
        return jsonify({"error": "indication_id + agency 필요", "code": "MISSING_FIELDS"}), 400
    valid_agencies = {"FDA", "EMA", "MHRA", "PMDA", "TGA", "MFDS"}
    if agency not in valid_agencies:
        return jsonify({"error": f"agency 는 {sorted(valid_agencies)} 중 하나",
                        "code": "BAD_AGENCY"}), 400

    # indications_master row 존재 검증
    with db._connect() as conn:
        if not conn.execute(
            "SELECT 1 FROM indications_master WHERE indication_id = ?",
            (indication_id,),
        ).fetchone():
            return jsonify({"error": "indication_id not found",
                            "code": "NOT_FOUND"}), 404

    approval_date = (request.form.get("approval_date") or "").strip() or None
    label_excerpt = (request.form.get("label_excerpt") or "").strip() or None
    label_url = (request.form.get("label_url") or "").strip() or None
    notes = (request.form.get("notes") or "").strip() or None
    user_email = getattr(request, "user", {}).get("sub") if hasattr(request, "user") else None

    # 파일 저장 — 영구 보존, 캐시 디렉토리 분리
    target_dir = UPLOAD_DIR / agency
    target_dir.mkdir(parents=True, exist_ok=True)
    from datetime import datetime as _dt
    ts = _dt.now().strftime("%Y%m%d-%H%M%S")
    safe_indication = re.sub(r"[^a-zA-Z0-9_]+", "_", indication_id)[:80]
    target = target_dir / f"{safe_indication}_{agency}_{ts}.pdf"
    try:
        file.save(str(target))
        size = target.stat().st_size
    except Exception as e:
        logger.error("PDF 저장 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e), "code": "SAVE_FAIL"}), 500

    # DB 등록
    try:
        rec = {
            "indication_id":     indication_id,
            "agency":            agency,
            "file_path":         str(target.relative_to(BASE_DIR)),
            "original_filename": file.filename,
            "file_size":         size,
            "content_type":      file.content_type or "application/pdf",
            "approval_date":     approval_date,
            "label_excerpt":     label_excerpt,
            "label_url":         label_url,
            "notes":             notes,
            "uploaded_by":       user_email,
        }
        doc_id = db.insert_approval_document(rec)

        # indications_by_agency 동시 갱신 — label_url 에 우리 PDF 경로, approval_date 보존
        from datetime import datetime as _dt2
        with db._connect() as conn:
            existing = conn.execute(
                "SELECT label_excerpt, label_url, approval_date "
                "FROM indications_by_agency "
                "WHERE indication_id = ? AND agency = ?",
                (indication_id, agency),
            ).fetchone()
            new_label_url = f"/api/admin/approval-document/{doc_id}/file"
            new_approval = approval_date or (existing[2] if existing else None)
            new_excerpt = label_excerpt or (existing[0] if existing else None)
            conn.execute(
                """
                INSERT INTO indications_by_agency
                    (indication_id, agency, biomarker_label, combination_label,
                     approval_date, label_excerpt, label_full_text, label_url,
                     restriction_note, raw_source, fetched_at)
                VALUES (?, ?, NULL, NULL, ?, ?, NULL, ?, NULL,
                        json_object('source','admin_pdf_upload','doc_id',?), ?)
                ON CONFLICT(indication_id, agency) DO UPDATE SET
                    approval_date = COALESCE(excluded.approval_date, indications_by_agency.approval_date),
                    label_excerpt = COALESCE(excluded.label_excerpt, indications_by_agency.label_excerpt),
                    label_url     = excluded.label_url,
                    fetched_at    = excluded.fetched_at
                """,
                (indication_id, agency, new_approval, new_excerpt, new_label_url,
                 doc_id, _dt2.now().isoformat(timespec="seconds")),
            )
            conn.commit()

        return jsonify({"ok": True, "doc_id": doc_id, "file_path": rec["file_path"]})
    except Exception as e:
        # 파일은 남기되 DB 실패는 알림
        logger.error("approval_document insert 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e), "code": "DB_FAIL"}), 500


@app.get("/api/admin/approval-document")
@require_auth(role="admin")
def admin_approval_document_list():
    """업로드된 PDF 메타 리스트.
    GET /api/admin/approval-document?indication_id=X / agency=X / product=X
    """
    indication_id = (request.args.get("indication_id") or "").strip() or None
    agency = (request.args.get("agency") or "").strip().upper() or None
    product = (request.args.get("product") or "").strip().lower() or None
    try:
        items = db.list_approval_documents(indication_id, agency, product)
        return jsonify({"items": items})
    except Exception as e:
        logger.error("approval_document_list 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.get("/api/admin/approval-document/<int:doc_id>/file")
def admin_approval_document_file(doc_id: int):
    """업로드된 PDF 다운로드/표시. 인증 무관 (임베드 view 위해).

    UI 의 iframe / link 에서 직접 호출 가능.
    """
    doc = db.get_approval_document(doc_id)
    if not doc:
        return jsonify({"error": "not found"}), 404
    p = BASE_DIR / doc["file_path"]
    if not p.exists():
        return jsonify({"error": "file missing on disk", "expected": str(p)}), 404
    from flask import send_file as _send_file
    return _send_file(
        str(p),
        mimetype=doc.get("content_type") or "application/pdf",
        as_attachment=False,
        download_name=doc.get("original_filename") or p.name,
    )


@app.delete("/api/admin/approval-document/<int:doc_id>")
@require_auth(role="admin")
def admin_approval_document_delete(doc_id: int):
    """row + file 삭제."""
    file_path = db.delete_approval_document(doc_id)
    if file_path is None:
        return jsonify({"error": "not found"}), 404
    p = BASE_DIR / file_path
    try:
        if p.exists():
            p.unlink()
    except OSError as e:
        logger.warning("PDF unlink 실패 (DB row 는 이미 삭제됨): %s", e)
    return jsonify({"ok": True, "deleted_id": doc_id})


# ──────────────────────────────────────────────────────────────────────────────
# FDA 자동 sync — admin 트리거 (LLM/스크레이퍼 기반 build)
# ──────────────────────────────────────────────────────────────────────────────

@app.post("/api/admin/fda-sync")
@require_auth(role="admin")
def admin_fda_sync():
    """FDA 적응증 자동 build (ForeignApprovalAgent.build 의 FDA agency 만).

    body:
      {drug: 'pembrolizumab', product_slug: 'keytruda', wipe: false}
    """
    body = request.get_json(silent=True) or {}
    drug = (body.get("drug") or "").strip()
    product_slug = (body.get("product_slug") or "").strip().lower()
    wipe = bool(body.get("wipe", False))
    if not drug or not product_slug:
        return jsonify({"error": "drug + product_slug 필요"}), 400

    try:
        from agents.foreign_approval import ForeignApprovalAgent
        agent = ForeignApprovalAgent(db_path=db.db_path)
        summary = agent.build(
            drug=drug,
            product_slug=product_slug,
            brand_slug=product_slug,
            agencies=["FDA"],
            wipe=wipe,
        )
        # AgencyBuildResult dataclass 를 dict 화
        result = {
            "product_slug": product_slug,
            "drug":         drug,
            "agencies":     [{"agency": a.agency, "ok": a.ok, "skipped": a.skipped,
                              "errors": a.errors} for a in summary.agencies],
        }
        return jsonify(result)
    except Exception as e:
        logger.error("fda_sync 실패: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# 헬스체크
# ──────────────────────────────────────────────────────────────────────────────

@app.get("/api/health")
def health():
    return jsonify({"status": "ok", "available_countries": AVAILABLE_COUNTRIES})


if __name__ == "__main__":
    logger.info("대쉬보드 API 서버 시작: http://127.0.0.1:5001")
    app.run(host="127.0.0.1", port=5001, debug=False)
